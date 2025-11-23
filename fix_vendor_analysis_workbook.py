#!/usr/bin/env python3
"""
Fix the VGM Vendor Analysis Workbook
- Corrects formula column placement (should be in L-P, not O-S)
- Fixes formula references to use correct columns
- Preserves all data and formatting
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

print("="*80)
print("FIXING VGM VENDOR ANALYSIS WORKBOOK")
print("="*80)

# Backup original file
original = "Holistic_Medical_VGM_Vendor_Analysis_COMPLETE_ALL_TIERS_2025-11-22.xlsx"
backup = f"Holistic_Medical_VGM_Vendor_Analysis_COMPLETE_ALL_TIERS_2025-11-22_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

print(f"\nBacking up original file...")
shutil.copy(original, backup)
print(f"  ✓ Backup saved: {backup}")

# Load workbook
print(f"\nLoading workbook...")
wb = openpyxl.load_workbook(original)
ws = wb['Inventory Analysis']

print(f"  ✓ Loaded {len(wb.sheetnames)} sheets")
print(f"  ✓ Inventory Analysis: {ws.max_row} rows, {ws.max_column} columns")

# First, clear wrong formulas from columns O-S (15-19)
print(f"\nClearing incorrect formulas from columns O-S...")
for row in range(2, ws.max_row + 1):
    ws[f'O{row}'].value = None  # Medicare Revenue
    ws[f'P{row}'].value = None  # Profit Margin %
    ws[f'Q{row}'].value = None  # Priority (should be value, not formula)
    ws[f'R{row}'].value = None  # Source (should be value, not formula)
    ws[f'S{row}'].value = None  # Customer (should be value, not formula)

print(f"  ✓ Cleared formulas from {ws.max_row - 1} rows")

# Now add CORRECT formulas to columns L-P
print(f"\nAdding correct formulas to columns L-P...")

for row in range(2, ws.max_row + 1):
    # L: Best Vendor - returns the name of vendor with lowest cost
    # Logic: IF(M=G, F, IF(M=I, H, J))
    # If Best Unit Cost (M) equals Vendor A Cost (G), return Vendor A Name (F)
    # Else if Best Unit Cost equals Vendor B Cost (I), return Vendor B Name (H)
    # Else return Vendor C Name (J)
    ws[f'L{row}'] = f'=IF(M{row}=G{row},F{row},IF(M{row}=I{row},H{row},J{row}))'

    # M: Best Unit Cost - minimum of three vendor costs
    # =MIN(G, I, K) where G=Vendor A Cost, I=Vendor B Cost, K=Vendor C Cost
    ws[f'M{row}'] = f'=MIN(G{row},I{row},K{row})'

    # N: Line Total Cost - Quantity * Best Unit Cost
    # =D*M where D=Quantity, M=Best Unit Cost
    ws[f'N{row}'] = f'=D{row}*M{row}'

    # O: Medicare Revenue - Quantity * Medicare Rate (if rate exists)
    # =IF(ISBLANK(E), 0, D*E) where E=Medicare Allowable Rate, D=Quantity
    ws[f'O{row}'] = f'=IF(ISBLANK(E{row}),0,D{row}*E{row})'

    # P: Profit Margin % - (Revenue - Cost) / Revenue
    # =IF(O=0, 0, (O-N)/O) where O=Medicare Revenue, N=Line Total Cost
    ws[f'P{row}'] = f'=IF(O{row}=0,0,(O{row}-N{row})/O{row})'

print(f"  ✓ Added formulas to {ws.max_row - 1} rows")

# Now restore the static data values for Priority, Source, and Customer from source CSV
print(f"\nRestoring static data values for Priority, Source, and Customer...")

import pandas as pd

# Load source data
df = pd.read_csv("/home/user/InitialInventoryScratch/MASTER_INVENTORY_PLAN_COMPLETE_ALL_TIERS.csv")
df = df.sort_values('Priority_Score', ascending=False).reset_index(drop=True)

# Verify row count matches
if len(df) != ws.max_row - 1:
    print(f"  ⚠️  WARNING: CSV has {len(df)} rows, worksheet has {ws.max_row - 1} data rows")
else:
    print(f"  ✓ Row counts match: {len(df)} rows")

# Restore values
for idx, row_data in df.iterrows():
    row_num = idx + 2  # Excel rows start at 2 (after header)

    ws[f'Q{row_num}'] = int(row_data['Priority_Score'])  # Priority
    ws[f'R{row_num}'] = row_data['Source']               # Source
    ws[f'S{row_num}'] = row_data['Customers'] if pd.notna(row_data['Customers']) else ""  # Customer

print(f"  ✓ Restored Priority, Source, and Customer values")

# Update conditional formatting for Profit Margin % column (now in P, not S)
print(f"\nUpdating conditional formatting...")

# Remove old conditional formatting
ws.conditional_formatting._cf_rules = {}

# Add new conditional formatting to column P
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

ws.conditional_formatting.add(f'P2:P{ws.max_row}',
    CellIsRule(operator='greaterThan', formula=['0.3'], stopIfTrue=True, fill=green_fill))
ws.conditional_formatting.add(f'P2:P{ws.max_row}',
    CellIsRule(operator='between', formula=['0.1', '0.3'], stopIfTrue=True, fill=yellow_fill))
ws.conditional_formatting.add(f'P2:P{ws.max_row}',
    CellIsRule(operator='lessThan', formula=['0.1'], stopIfTrue=True, fill=red_fill))

print(f"  ✓ Applied conditional formatting to column P (Profit Margin %)")

# Fix number formatting for all currency and percentage columns
print(f"\nApplying number formatting...")

for row in range(2, ws.max_row + 1):
    ws[f'E{row}'].number_format = '$#,##0.00'  # Medicare Allowable Rate
    ws[f'G{row}'].number_format = '$#,##0.00'  # Vendor A Unit Cost
    ws[f'I{row}'].number_format = '$#,##0.00'  # Vendor B Unit Cost
    ws[f'K{row}'].number_format = '$#,##0.00'  # Vendor C Unit Cost
    ws[f'M{row}'].number_format = '$#,##0.00'  # Best Unit Cost
    ws[f'N{row}'].number_format = '$#,##0.00'  # Line Total Cost
    ws[f'O{row}'].number_format = '$#,##0.00'  # Medicare Revenue
    ws[f'P{row}'].number_format = '0.0%'       # Profit Margin %

print(f"  ✓ Applied number formatting")

# Save corrected workbook
output_file = original
wb.save(output_file)

print(f"\n{'='*80}")
print(f"✅ WORKBOOK FIXED")
print(f"{'='*80}")
print(f"  File: {output_file}")
print(f"  Backup: {backup}")
print(f"  Total rows: {ws.max_row - 1}")
print(f"\nCorrected Formulas:")
print(f"  L (Best Vendor): =IF(M=G,F,IF(M=I,H,J))")
print(f"  M (Best Unit Cost): =MIN(G,I,K)")
print(f"  N (Line Total Cost): =D*M")
print(f"  O (Medicare Revenue): =IF(ISBLANK(E),0,D*E)")
print(f"  P (Profit Margin %): =IF(O=0,0,(O-N)/O)")
print(f"\nStatic Data Columns:")
print(f"  Q (Priority): Restored from CSV")
print(f"  R (Source): Restored from CSV")
print(f"  S (Customer): Restored from CSV")
print(f"{'='*80}")
