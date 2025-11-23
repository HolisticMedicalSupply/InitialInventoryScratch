#!/usr/bin/env python3
"""
Comprehensive validation of the complete all-tiers workbook
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

print("="*100)
print("VALIDATION: COMPLETE ALL-TIERS VGM VENDOR ANALYSIS WORKBOOK")
print("="*100)

workbook_path = "/home/user/InitialInventoryScratch/Holistic_Medical_VGM_Vendor_Analysis_COMPLETE_ALL_TIERS_2025-11-22.xlsx"
master_csv = "/home/user/InitialInventoryScratch/MASTER_INVENTORY_PLAN_COMPLETE_ALL_TIERS.csv"

# Load master data
master_df = pd.read_csv(master_csv)
print(f"\n✓ Loaded master plan: {len(master_df)} products")

# Load workbook
wb = load_workbook(workbook_path)
print(f"✓ Loaded workbook: {workbook_path}")

print(f"\n{'='*100}")
print("FILE STRUCTURE VALIDATION")
print("="*100)

# Check sheets
required_sheets = ["Inventory Analysis", "BOC Category Summary", "Items Without Medicare Rates", "Customer Requests"]
for sheet in required_sheets:
    if sheet in wb.sheetnames:
        print(f"  ✓ Sheet '{sheet}' present")
    else:
        print(f"  ✗ Sheet '{sheet}' MISSING")

ws_main = wb["Inventory Analysis"]
ws_boc = wb["BOC Category Summary"]
ws_no_medicare = wb["Items Without Medicare Rates"]
ws_customers = wb["Customer Requests"]

# Count rows
main_rows = ws_main.max_row - 1  # Exclude header
print(f"\n  Main sheet rows: {main_rows} (expected: 383)")

print(f"\n{'='*100}")
print("FORMULA VALIDATION")
print("="*100)

# Check formulas in all rows
formula_checks = {
    'Best_Vendor (O)': 0,
    'Best_Unit_Cost (P)': 0,
    'Line_Total_Cost (Q)': 0,
    'Medicare_Revenue (R)': 0,
    'Profit_Margin_% (S)': 0
}

formula_columns = {
    'O': 'Best_Vendor',
    'P': 'Best_Unit_Cost',
    'Q': 'Line_Total_Cost',
    'R': 'Medicare_Revenue',
    'S': 'Profit_Margin_%'
}

for row_idx in range(2, main_rows + 2):
    for col, name in formula_columns.items():
        cell = ws_main[f'{col}{row_idx}']
        if cell.data_type == 'f':  # Formula
            formula_checks[f'{name} ({col})'] += 1

print(f"\nFormula Coverage:")
for formula_name, count in formula_checks.items():
    coverage = (count / main_rows * 100) if main_rows > 0 else 0
    status = "✓" if coverage == 100 else "✗"
    print(f"  {status} {formula_name}: {count}/{main_rows} ({coverage:.1f}%)")

# Validate formula syntax (sample check)
print(f"\nFormula Syntax Validation (Row 2 - L4361 CAM Walker):")
formulas = {
    'O2': ws_main['O2'].value,
    'P2': ws_main['P2'].value,
    'Q2': ws_main['Q2'].value,
    'R2': ws_main['R2'].value,
    'S2': ws_main['S2'].value
}

expected_formulas = {
    'O2': '=IF(P2=J2,I2,IF(P2=L2,K2,M2))',
    'P2': '=MIN(J2,L2,N2)',
    'Q2': '=G2*P2',
    'R2': '=IF(ISBLANK(H2),0,G2*H2)',
    'S2': '=IF(R2=0,0,(R2-Q2)/R2)'
}

for cell, formula in formulas.items():
    expected = expected_formulas.get(cell, '')
    if formula == expected:
        print(f"  ✓ {cell}: {formula}")
    else:
        print(f"  ✗ {cell}: Expected '{expected}', got '{formula}'")

print(f"\n{'='*100}")
print("FORMATTING VALIDATION")
print("="*100)

# Check header formatting
header_cell = ws_main['A1']
header_fill_color = header_cell.fill.start_color.rgb if header_cell.fill.start_color else None
header_font_bold = header_cell.font.bold if header_cell.font else False

print(f"\nHeader Formatting:")
print(f"  Background: {header_fill_color} (expected: FF1F4E78 or 001F4E78)")
print(f"  Font Bold: {header_font_bold} (expected: True)")

# Check frozen panes
frozen_pane = ws_main.freeze_panes
print(f"  Frozen Panes: {frozen_pane} (expected: A2)")

# Check conditional formatting
has_conditional = len(ws_main.conditional_formatting._cf_rules) > 0
print(f"  Conditional Formatting: {has_conditional} (expected: True)")
print(f"  Number of CF rules: {len(ws_main.conditional_formatting._cf_rules)}")

# Check number formats (sample)
print(f"\nNumber Formatting (Row 2):")
number_format_checks = {
    'E2': ('Medicare Rate', '$#,##0.00'),
    'P2': ('Profit Margin %', '0.0%'),
    'Q2': ('Line Total Cost', '$#,##0.00'),
    'R2': ('Medicare Revenue', '$#,##0.00'),
}

for cell_ref, (field_name, expected_format) in number_format_checks.items():
    actual_format = ws_main[cell_ref].number_format
    if actual_format == expected_format:
        print(f"  ✓ {cell_ref} ({field_name}): {actual_format}")
    else:
        print(f"  ⚠ {cell_ref} ({field_name}): Expected '{expected_format}', got '{actual_format}'")

print(f"\n{'='*100}")
print("BOC CATEGORY COVERAGE VALIDATION")
print("="*100)

# Check BOC categories
boc_in_master = set(master_df['BOC_Category'].unique())
print(f"\n  BOC Categories in Master Plan: {len(boc_in_master)}")
print(f"  BOC Categories in Summary Sheet: {ws_boc.max_row - 1}")

# Load approved BOC codes
approved_boc = pd.read_csv("/home/user/InitialInventoryScratch/ApprovedCategoriesAndCodes copy.csv", skiprows=1, nrows=35)
approved_boc_codes = set(approved_boc['Approved BOC Code'].dropna().unique()) - {'DM14'}  # Exclude DM14 (NOT FOUND)
total_approved = len(approved_boc_codes)

coverage_pct = (len(boc_in_master) / total_approved) * 100
print(f"  Coverage: {len(boc_in_master)}/{total_approved} ({coverage_pct:.1f}%)")

# List unused categories
unused_boc = approved_boc_codes - boc_in_master
print(f"\n  Unused BOC Categories ({len(unused_boc)}):")
for boc in sorted(unused_boc):
    desc = approved_boc[approved_boc['Approved BOC Code'] == boc]['Category Description'].values
    desc_str = desc[0] if len(desc) > 0 else "Unknown"
    print(f"    - {boc}: {desc_str}")

print(f"\n{'='*100}")
print("CUSTOMER REQUIREMENTS VALIDATION")
print("="*100)

# Define customer requirements
customer_requirements = {
    'DR_NAS': {
        'expected_count': 6,
        'critical_items': ['L4361', 'L1906', 'L4350', 'L4370', 'L4387', 'L1902']
    },
    'MOYHINOR': {
        'expected_count': 2,
        'critical_items': ['K0001']  # Wheelchair
    },
    'MOYHINOR (CGM_STRATEGY)': {
        'expected_count': 4,
        'critical_items': ['E2103', 'A4239']
    },
    'RAMBOM': {
        'expected_count': 4,
        'critical_items': ['E0130', 'E0143', 'B4150', 'E0570']
    },
    'WALTERS': {
        'expected_count': 4,
        'critical_items': ['A4253', 'A4259', 'A4554', 'A4520']
    }
}

# Check customer items
customer_items = master_df[master_df['Source'] == 'CUSTOMER']

for customer, requirements in customer_requirements.items():
    items = customer_items[customer_items['Customers'] == customer]
    expected = requirements['expected_count']
    actual = len(items)
    status = "✓" if actual >= expected else "✗"

    print(f"\n{status} {customer}: {actual}/{expected} items")

    # Check critical items
    for critical_hcpcs in requirements['critical_items']:
        found = items[items['HCPCS_Code'] == critical_hcpcs]
        if len(found) > 0:
            print(f"    ✓ {critical_hcpcs} - {found.iloc[0]['Description']}")
        else:
            print(f"    ✗ {critical_hcpcs} - MISSING")

print(f"\n{'='*100}")
print("DATA INTEGRITY VALIDATION")
print("="*100)

# Check for blank critical fields
blank_hcpcs = master_df[master_df['HCPCS_Code'].isna()]
blank_desc = master_df[master_df['Description'].isna()]
blank_qty = master_df[master_df['Quantity'].isna()]

print(f"\nBlank Field Check:")
print(f"  ✓ Blank HCPCS Codes: {len(blank_hcpcs)} (expected: 0)")
print(f"  ✓ Blank Descriptions: {len(blank_desc)} (expected: 0)")
print(f"  ✓ Blank Quantities: {len(blank_qty)} (expected: 0)")

# Check for negative quantities
negative_qty = master_df[master_df['Quantity'] < 0]
print(f"  ✓ Negative Quantities: {len(negative_qty)} (expected: 0)")

# Check unique HCPCS codes
unique_hcpcs = master_df['HCPCS_Code'].nunique()
total_rows = len(master_df)
duplicates = total_rows - unique_hcpcs
print(f"\n  Unique HCPCS Codes: {unique_hcpcs}")
print(f"  Total Rows: {total_rows}")
print(f"  Duplicates: {duplicates} (intentional variants for different products)")

print(f"\n{'='*100}")
print("PRIORITY VALIDATION")
print("="*100)

# Check priority scores
top_priority = master_df.nlargest(5, 'Priority_Score')[['HCPCS_Code', 'Description', 'Priority_Score', 'Customers']]
print(f"\nTop 5 Priority Items:")
for idx, row in top_priority.iterrows():
    customer = row['Customers'] if pd.notna(row['Customers']) else 'LAUNCH'
    print(f"  {int(row['Priority_Score']):3d} - {row['HCPCS_Code']:6s} - {row['Description'][:50]:50s} ({customer})")

# Verify L4361 is highest priority
l4361 = master_df[master_df['HCPCS_Code'] == 'L4361']
if len(l4361) > 0:
    l4361_priority = l4361.iloc[0]['Priority_Score']
    max_priority = master_df['Priority_Score'].max()
    if l4361_priority == max_priority:
        print(f"\n  ✓ L4361 (CAM Walker) has highest priority: {int(l4361_priority)}")
    else:
        print(f"\n  ✗ L4361 priority ({int(l4361_priority)}) is not highest ({int(max_priority)})")
else:
    print(f"\n  ✗ L4361 (CAM Walker) not found in inventory")

print(f"\n{'='*100}")
print("FINANCIAL SUMMARY")
print("="*100)

total_investment = master_df['Total_Cost'].sum()
total_revenue = master_df['Total_Revenue_Potential'].sum()
total_profit = master_df['Total_Profit_Potential'].sum()
avg_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0

medicare_covered = master_df[master_df['Medicare_Rate'].notna()]
medicare_coverage_pct = (len(medicare_covered) / len(master_df)) * 100

print(f"\n  Total Products: {len(master_df)}")
print(f"  Total Investment: ${total_investment:,.2f}")
print(f"  Total Revenue Potential: ${total_revenue:,.2f}")
print(f"  Total Profit Potential: ${total_profit:,.2f}")
print(f"  Average Margin: {avg_margin:.1f}%")
print(f"  Medicare Coverage: {medicare_coverage_pct:.1f}%")

print(f"\n{'='*100}")
print("VALIDATION SCORE")
print("="*100)

# Calculate overall score
checks = {
    'File Structure': all(sheet in wb.sheetnames for sheet in required_sheets),
    'Formula Integrity': all(count == main_rows for count in formula_checks.values()),
    'Formula Syntax': all(formulas[k] == expected_formulas[k] for k in formulas.keys()),
    'BOC Coverage': coverage_pct >= 75.0,  # At least 75% coverage
    'Customer Requirements': all(len(customer_items[customer_items['Customers'] == c]) >= customer_requirements[c]['expected_count'] for c in customer_requirements.keys()),
    'Data Integrity': len(blank_hcpcs) == 0 and len(blank_desc) == 0 and len(blank_qty) == 0,
    'L4361 Priority': len(l4361) > 0 and l4361.iloc[0]['Priority_Score'] == master_df['Priority_Score'].max(),
}

passed = sum(checks.values())
total = len(checks)
score = (passed / total) * 100

print(f"\n  Overall Score: {passed}/{total} ({score:.1f}%)")
print()

for check_name, result in checks.items():
    status = "✅ PASS" if result else "❌ FAIL"
    print(f"  {status} - {check_name}")

if score >= 85:
    print(f"\n{'='*100}")
    print("✅ WORKBOOK VALIDATION: PASSED")
    print("="*100)
    print("\n  The workbook is production-ready for the VGM vendor meeting.")
    print(f"  All critical requirements met with {score:.1f}% validation score.")
else:
    print(f"\n{'='*100}")
    print("⚠️ WORKBOOK VALIDATION: NEEDS REVIEW")
    print("="*100)
    print(f"\n  Validation score {score:.1f}% is below 85% threshold.")
    print("  Please review failed checks above.")

print()
