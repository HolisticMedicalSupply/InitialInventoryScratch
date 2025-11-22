#!/usr/bin/env python3
"""
Comprehensive validation of final VGM vendor analysis workbook
- Validate all formulas work correctly
- Verify calculations are accurate
- Check formatting is professional
- Confirm all customer requirements met
- Ensure data integrity
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os

print("="*100)
print("COMPREHENSIVE WORKBOOK VALIDATION")
print("="*100)

workbook_path = "/home/user/InitialInventoryScratch/Holistic_Medical_VGM_Vendor_Analysis_FINAL_2025-11-22.xlsx"

# ============================================================================
# VALIDATION 1: FILE STRUCTURE
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 1: FILE STRUCTURE")
print("="*100)

if not os.path.exists(workbook_path):
    print("❌ FAIL: Workbook file not found")
    exit(1)

file_size = os.path.getsize(workbook_path)
print(f"\n✓ File exists: {workbook_path}")
print(f"✓ File size: {file_size:,} bytes")

wb = load_workbook(workbook_path, data_only=False)  # Load with formulas
wb_data = load_workbook(workbook_path, data_only=True)  # Load with values

expected_sheets = ['Inventory Analysis', 'BOC Category Summary', 'Items Without Medicare Rates', 'Customer Requests']
actual_sheets = wb.sheetnames

print(f"\n✓ Expected sheets: {len(expected_sheets)}")
print(f"✓ Actual sheets: {len(actual_sheets)}")

all_sheets_present = all(sheet in actual_sheets for sheet in expected_sheets)
if all_sheets_present:
    print(f"✅ PASS: All required sheets present")
    for sheet in expected_sheets:
        print(f"    ✓ {sheet}")
else:
    print(f"❌ FAIL: Missing sheets")
    missing = set(expected_sheets) - set(actual_sheets)
    print(f"    Missing: {missing}")

# ============================================================================
# VALIDATION 2: FORMULA INTEGRITY
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 2: FORMULA INTEGRITY")
print("="*100)

ws_formulas = wb['Inventory Analysis']
total_rows = ws_formulas.max_row - 1  # Exclude header

print(f"\nTotal product rows: {total_rows}")

# Check formulas in specific columns
formula_columns = {
    'O': 'Best_Vendor',
    'P': 'Best_Unit_Cost',
    'Q': 'Line_Total_Cost',
    'R': 'Medicare_Revenue_Potential',
    'S': 'Profit_Margin_%'
}

formula_check_results = {}

for col, name in formula_columns.items():
    # Check row 2 (first data row)
    cell = f'{col}2'
    cell_value = ws_formulas[cell].value

    if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
        formula_check_results[name] = 'PASS'
        print(f"✅ {name} ({col}2): {cell_value}")
    else:
        formula_check_results[name] = 'FAIL'
        print(f"❌ {name} ({col}2): No formula found - value is {cell_value}")

# Check all rows have formulas
print(f"\nChecking formulas across all {total_rows} rows...")
formula_counts = {col: 0 for col in formula_columns.keys()}

for row_idx in range(2, ws_formulas.max_row + 1):
    for col in formula_columns.keys():
        cell_value = ws_formulas[f'{col}{row_idx}'].value
        if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
            formula_counts[col] += 1

print(f"\nFormula coverage:")
for col, name in formula_columns.items():
    count = formula_counts[col]
    pct = (count / total_rows) * 100 if total_rows > 0 else 0
    status = "✅" if count == total_rows else "⚠️"
    print(f"  {status} {name}: {count}/{total_rows} ({pct:.1f}%)")

all_formulas_ok = all(count == total_rows for count in formula_counts.values())
if all_formulas_ok:
    print(f"\n✅ PASS: All formula columns have formulas in all rows")
else:
    print(f"\n⚠️ WARNING: Some rows missing formulas")

# ============================================================================
# VALIDATION 3: CALCULATION ACCURACY
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 3: CALCULATION ACCURACY")
print("="*100)

# Load data values (after formulas evaluated)
ws_data = wb_data['Inventory Analysis']

print(f"\nValidating calculations for sample rows...")

calculation_errors = []

for row_idx in [2, 3, 10, 50, 100, 200, total_rows + 1]:  # Sample rows
    if row_idx > ws_data.max_row:
        continue

    # Get values
    qty = ws_data[f'G{row_idx}'].value
    vendor_a_cost = ws_data[f'J{row_idx}'].value
    vendor_b_cost = ws_data[f'L{row_idx}'].value
    vendor_c_cost = ws_data[f'N{row_idx}'].value
    best_cost = ws_data[f'P{row_idx}'].value
    line_total = ws_data[f'Q{row_idx}'].value
    medicare_rate = ws_data[f'H{row_idx}'].value
    medicare_revenue = ws_data[f'R{row_idx}'].value
    margin_pct = ws_data[f'S{row_idx}'].value
    hcpcs = ws_data[f'C{row_idx}'].value

    # Validate Best_Unit_Cost = MIN(vendor costs)
    vendor_costs = [c for c in [vendor_a_cost, vendor_b_cost, vendor_c_cost] if c is not None]
    expected_best_cost = min(vendor_costs) if vendor_costs else None

    if expected_best_cost is not None and best_cost is not None:
        if abs(best_cost - expected_best_cost) > 0.01:
            calculation_errors.append(f"Row {row_idx} ({hcpcs}): Best cost mismatch - Expected {expected_best_cost}, got {best_cost}")

    # Validate Line_Total_Cost = Quantity * Best_Unit_Cost
    if qty and best_cost:
        expected_line_total = qty * best_cost
        if line_total and abs(line_total - expected_line_total) > 0.01:
            calculation_errors.append(f"Row {row_idx} ({hcpcs}): Line total mismatch - Expected {expected_line_total:.2f}, got {line_total:.2f}")

    # Validate Medicare_Revenue_Potential = Quantity * Medicare_Rate
    if medicare_rate and qty:
        expected_medicare_revenue = qty * medicare_rate
        if medicare_revenue and abs(medicare_revenue - expected_medicare_revenue) > 0.01:
            calculation_errors.append(f"Row {row_idx} ({hcpcs}): Medicare revenue mismatch - Expected {expected_medicare_revenue:.2f}, got {medicare_revenue:.2f}")

    # Validate Profit_Margin_% = (Revenue - Cost) / Revenue
    if medicare_revenue and medicare_revenue > 0 and line_total:
        expected_margin = (medicare_revenue - line_total) / medicare_revenue
        if margin_pct is not None and abs(margin_pct - expected_margin) > 0.001:  # 0.1% tolerance
            calculation_errors.append(f"Row {row_idx} ({hcpcs}): Margin % mismatch - Expected {expected_margin:.4f}, got {margin_pct:.4f}")

if calculation_errors:
    print(f"❌ FAIL: {len(calculation_errors)} calculation errors found")
    for error in calculation_errors[:10]:  # Show first 10
        print(f"    {error}")
else:
    print(f"✅ PASS: All sampled calculations accurate")

# ============================================================================
# VALIDATION 4: FORMATTING CHECK
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 4: FORMATTING CHECK")
print("="*100)

ws_format = wb['Inventory Analysis']

# Check header formatting
header_cell = ws_format['A1']
header_font_ok = header_cell.font.bold and header_cell.font.color.rgb == 'FFFFFFFF'
header_fill_ok = header_cell.fill.start_color.rgb == '001F4E78'

print(f"\nHeader formatting:")
print(f"  ✓ Font: {'Bold White' if header_font_ok else 'INCORRECT'}")
print(f"  ✓ Background: {'Dark Blue (#1F4E78)' if header_fill_ok else 'INCORRECT'}")

if header_font_ok and header_fill_ok:
    print(f"✅ PASS: Header formatting correct")
else:
    print(f"❌ FAIL: Header formatting incorrect")

# Check freeze panes
freeze_panes_ok = ws_format.freeze_panes == 'A2'
print(f"\nFreeze panes: {ws_format.freeze_panes}")
if freeze_panes_ok:
    print(f"✅ PASS: Freeze panes set correctly (A2)")
else:
    print(f"⚠️ WARNING: Freeze panes not set or incorrect")

# Check conditional formatting
cond_format_rules = ws_format.conditional_formatting._cf_rules
has_cond_format = len(cond_format_rules) > 0

print(f"\nConditional formatting:")
print(f"  Rules defined: {len(cond_format_rules)}")
if has_cond_format:
    print(f"✅ PASS: Conditional formatting present")
    for range_str, rules in cond_format_rules.items():
        print(f"    Range: {range_str}, Rules: {len(rules)}")
else:
    print(f"⚠️ WARNING: No conditional formatting found")

# Check number formatting
print(f"\nNumber formatting (sample cells):")
format_checks = [
    ('G2', 'Quantity', '0'),
    ('H2', 'Medicare Rate', '$#,##0.00'),
    ('J2', 'Vendor A Cost', '$#,##0.00'),
    ('P2', 'Best Cost', '$#,##0.00'),
    ('Q2', 'Line Total', '$#,##0.00'),
    ('R2', 'Revenue', '$#,##0.00'),
    ('S2', 'Margin %', '0.0%')
]

format_ok_count = 0
for cell, name, expected_format in format_checks:
    actual_format = ws_format[cell].number_format
    # Normalize format strings for comparison
    matches = actual_format == expected_format or \
              (expected_format == '$#,##0.00' and '#,##0.00' in actual_format) or \
              (expected_format == '0.0%' and '%' in actual_format)

    status = "✓" if matches else "✗"
    if matches:
        format_ok_count += 1
    print(f"  {status} {name} ({cell}): {actual_format}")

format_pct = (format_ok_count / len(format_checks)) * 100
if format_ok_count == len(format_checks):
    print(f"✅ PASS: All number formats correct ({format_pct:.0f}%)")
else:
    print(f"⚠️ WARNING: {format_ok_count}/{len(format_checks)} formats correct ({format_pct:.0f}%)")

# Column widths
print(f"\nColumn widths (sample):")
width_checks = [('A', 15), ('D', 60), ('F', 30), ('T', 50)]
for col, expected_width in width_checks:
    actual_width = ws_format.column_dimensions[col].width
    status = "✓" if abs(actual_width - expected_width) < 5 else "✗"  # 5-unit tolerance
    print(f"  {status} Column {col}: {actual_width:.1f} (expected ~{expected_width})")

# ============================================================================
# VALIDATION 5: CUSTOMER REQUIREMENTS
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 5: CUSTOMER REQUIREMENTS")
print("="*100)

# Load customer requests sheet
ws_customer_data = wb_data['Customer Requests']

print(f"\nCustomer Request Items: {ws_customer_data.max_row - 1}")

# Expected customer items
expected_customer_items = {
    'DR_NAS': ['L1902', 'L1906', 'L4361', 'L4350', 'L4370', 'L4387'],  # 6 ankle products
    'MOYHINOR': ['E0607', 'K0001'],  # Glucose monitor, kids wheelchair
    'MOYHINOR (CGM_STRATEGY)': ['E2103', 'A4239'],  # CGM systems
    'RAMBOM': ['B4150', 'E0143', 'E0130', 'E0570'],  # Formula, rollator, walker, nebulizer
    'WALTERS': ['A4253', 'A4259', 'A4554', 'A4520']  # Glucose strips, lancets, chux, diapers
}

# Extract HCPCS codes from customer sheet
customer_items_found = {}

for row_idx in range(2, ws_customer_data.max_row + 1):
    hcpcs = ws_customer_data[f'C{row_idx}'].value
    customers = ws_customer_data[f'F{row_idx}'].value

    if customers:
        if customers not in customer_items_found:
            customer_items_found[customers] = []
        customer_items_found[customers].append(hcpcs)

print(f"\nCustomer Coverage:")
total_expected = 0
total_found = 0

for customer, expected_codes in expected_customer_items.items():
    found_codes = customer_items_found.get(customer, [])
    missing_codes = set(expected_codes) - set(found_codes)

    total_expected += len(expected_codes)
    total_found += len(found_codes)

    coverage = len(found_codes) / len(expected_codes) * 100 if expected_codes else 0
    status = "✅" if coverage == 100 else "⚠️"

    print(f"\n{status} {customer}: {len(found_codes)}/{len(expected_codes)} ({coverage:.0f}%)")
    for code in expected_codes:
        found = "✓" if code in found_codes else "✗"
        special = ""
        if code == 'L4361':
            special = " (PRIMARY REQUEST)"
        print(f"    {found} {code}{special}")

    if missing_codes:
        print(f"    Missing: {missing_codes}")

overall_coverage = (total_found / total_expected) * 100 if total_expected > 0 else 0
print(f"\n✅ Overall Customer Coverage: {total_found}/{total_expected} ({overall_coverage:.1f}%)")

# ============================================================================
# VALIDATION 6: DATA INTEGRITY
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 6: DATA INTEGRITY")
print("="*100)

ws_data_check = wb_data['Inventory Analysis']

# Check for blank critical fields
critical_columns = {
    'C': 'HCPCS_Code',
    'D': 'Description',
    'G': 'Quantity'
}

blank_checks = {col: 0 for col in critical_columns.keys()}

for row_idx in range(2, ws_data_check.max_row + 1):
    for col in critical_columns.keys():
        value = ws_data_check[f'{col}{row_idx}'].value
        if value is None or (isinstance(value, str) and value.strip() == ''):
            blank_checks[col] += 1

print(f"\nBlank field checks (critical columns):")
any_blanks = False
for col, name in critical_columns.items():
    count = blank_checks[col]
    status = "✓" if count == 0 else "✗"
    if count > 0:
        any_blanks = True
    print(f"  {status} {name}: {count} blank cells")

if not any_blanks:
    print(f"✅ PASS: No blank critical fields")
else:
    print(f"❌ FAIL: Found blank critical fields")

# Check for duplicate HCPCS codes
hcpcs_list = []
for row_idx in range(2, ws_data_check.max_row + 1):
    hcpcs = ws_data_check[f'C{row_idx}'].value
    if hcpcs:
        hcpcs_list.append(hcpcs)

unique_hcpcs = len(set(hcpcs_list))
total_hcpcs = len(hcpcs_list)

print(f"\nHCPCS Code uniqueness:")
print(f"  Total codes: {total_hcpcs}")
print(f"  Unique codes: {unique_hcpcs}")

if unique_hcpcs == total_hcpcs:
    print(f"✅ PASS: All HCPCS codes unique")
else:
    duplicates = total_hcpcs - unique_hcpcs
    print(f"⚠️ WARNING: {duplicates} duplicate HCPCS codes (may be intentional for different products)")

# Check quantity values are positive
negative_qty_count = 0
for row_idx in range(2, ws_data_check.max_row + 1):
    qty = ws_data_check[f'G{row_idx}'].value
    if qty is not None and qty <= 0:
        negative_qty_count += 1

if negative_qty_count == 0:
    print(f"\n✅ PASS: All quantities are positive")
else:
    print(f"\n❌ FAIL: {negative_qty_count} rows have zero or negative quantities")

# ============================================================================
# VALIDATION 7: BUDGET TOTALS
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 7: BUDGET TOTALS")
print("="*100)

# Calculate totals
total_investment = 0
total_revenue = 0

for row_idx in range(2, ws_data_check.max_row + 1):
    line_cost = ws_data_check[f'Q{row_idx}'].value
    line_revenue = ws_data_check[f'R{row_idx}'].value

    if line_cost:
        total_investment += line_cost
    if line_revenue:
        total_revenue += line_revenue

total_profit = total_revenue - total_investment

print(f"\nBudget Summary:")
print(f"  Total Investment: ${total_investment:,.2f}")
print(f"  Total Revenue Potential: ${total_revenue:,.2f}")
print(f"  Total Profit Potential: ${total_profit:,.2f}")

if total_revenue > 0:
    overall_margin = (total_profit / total_revenue) * 100
    print(f"  Overall Margin: {overall_margin:.1f}%")

expected_investment_min = 67000
expected_investment_max = 68000

if expected_investment_min <= total_investment <= expected_investment_max:
    print(f"\n✅ PASS: Total investment within expected range (${expected_investment_min:,} - ${expected_investment_max:,})")
else:
    print(f"\n⚠️ WARNING: Total investment outside expected range")

# ============================================================================
# VALIDATION 8: CRITICAL ITEMS CHECK
# ============================================================================
print("\n" + "="*100)
print("VALIDATION 8: CRITICAL ITEMS CHECK")
print("="*100)

# Check for critical items
critical_items = {
    'L4361': 'Dr. Nas CAM walker (PRIMARY REQUEST)',
    'L1902': 'Dr. Nas ankle gauntlet',
    'L1906': 'Dr. Nas multiligamentous support',
    'L4350': 'Dr. Nas ankle control',
    'L4370': 'Dr. Nas pneumatic splint',
    'L4387': 'Dr. Nas walking boot',
    'E2103': 'CGM Receiver (Dexcom/Libre)',
    'A4239': 'CGM Sensors (Dexcom/Libre)',
    'A4253': "Walter's glucose strips",
    'A4520': "Walter's adult diapers"
}

print(f"\nCritical items presence:")
all_critical_found = True

for hcpcs, description in critical_items.items():
    found = hcpcs in hcpcs_list
    status = "✓" if found else "✗"
    if not found:
        all_critical_found = False
    print(f"  {status} {hcpcs}: {description}")

if all_critical_found:
    print(f"\n✅ PASS: All critical items present")
else:
    print(f"\n❌ FAIL: Some critical items missing")

# ============================================================================
# FINAL SUMMARY
# ============================================================================
print("\n" + "="*100)
print("VALIDATION SUMMARY")
print("="*100)

validation_results = {
    'File Structure': all_sheets_present,
    'Formula Integrity': all_formulas_ok,
    'Calculation Accuracy': len(calculation_errors) == 0,
    'Formatting': header_font_ok and header_fill_ok and freeze_panes_ok,
    'Customer Requirements': overall_coverage >= 95,
    'Data Integrity': not any_blanks and negative_qty_count == 0,
    'Budget Totals': expected_investment_min <= total_investment <= expected_investment_max,
    'Critical Items': all_critical_found
}

print(f"\nValidation Results:")
pass_count = sum(1 for v in validation_results.values() if v)
total_checks = len(validation_results)

for check_name, result in validation_results.items():
    status = "✅ PASS" if result else "❌ FAIL"
    print(f"  {status}: {check_name}")

print(f"\nOverall Score: {pass_count}/{total_checks} ({pass_count/total_checks*100:.0f}%)")

if pass_count == total_checks:
    print(f"\n🎉 ✅ ALL VALIDATIONS PASSED - WORKBOOK IS PRODUCTION READY")
elif pass_count >= total_checks * 0.75:
    print(f"\n⚠️  MOST VALIDATIONS PASSED - Minor issues to address")
else:
    print(f"\n❌ CRITICAL ISSUES FOUND - Review and fix required")

print("="*100)
