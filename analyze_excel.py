#!/usr/bin/env python3
"""
Analyze existing Excel files to understand structure and data
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analyze_workbook(file_path):
    """Analyze an Excel workbook and print its structure"""
    print(f"\n{'='*80}")
    print(f"ANALYZING: {file_path}")
    print(f"{'='*80}")

    try:
        # Load workbook
        wb = load_workbook(file_path, read_only=False)
        print(f"\nSheet Names: {wb.sheetnames}")

        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'-'*80}")
            print(f"SHEET: {sheet_name}")
            print(f"{'-'*80}")

            # Read with pandas for better analysis
            df = pd.read_excel(file_path, sheet_name=sheet_name)

            print(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
            print(f"\nColumns: {list(df.columns)}")
            print(f"\nFirst 5 rows:")
            print(df.head().to_string())

            # Check for any non-null values to understand data completeness
            print(f"\nData completeness:")
            for col in df.columns:
                non_null = df[col].notna().sum()
                total = len(df)
                pct = (non_null/total*100) if total > 0 else 0
                print(f"  {col}: {non_null}/{total} ({pct:.1f}%)")

        wb.close()

    except Exception as e:
        print(f"ERROR: {e}")

# Analyze all Excel files
files_to_analyze = [
    "/home/user/InitialInventoryScratch/Medicare_Rates_Normalized_Structure_Validated.xlsx",
    "/home/user/InitialInventoryScratch/Holistic_Medical_Inventory_DETAILED(1).xlsx",
    "/home/user/InitialInventoryScratch/Claude Working Files & Convo log/Holistic_Medical_Inventory_DETAILED.xlsx"
]

for file_path in files_to_analyze:
    analyze_workbook(file_path)

print(f"\n{'='*80}")
print("ANALYSIS COMPLETE")
print(f"{'='*80}")
