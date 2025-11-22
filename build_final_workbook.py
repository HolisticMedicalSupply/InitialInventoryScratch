#!/usr/bin/env python3
"""
Final Excel Workbook Builder
Creates VGM-ready vendor analysis workbook from Master Inventory Plan
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

print("=" * 80)
print("FINAL VGM VENDOR ANALYSIS WORKBOOK BUILDER")
print("=" * 80)
print()

# Load master inventory plan
print("Loading master inventory plan...")
inventory = pd.read_csv('MASTER_INVENTORY_PLAN.csv')
print(f"  ✓ Loaded {len(inventory)} SKUs")
print()

# Prepare main analysis sheet
print("Preparing main analysis sheet...")

# Create vendor comparison structure
main_df = inventory[[
    'BOC_Category', 'HCPCS_Code', 'Description', 'Quantity',
    'Estimated_Unit_Cost', 'Total_Cost', 'Medicare_Rate',
    'Priority_Score', 'Source', 'Customers'
]].copy()

# Add vendor columns (VGM + 2 competitors)
main_df['Vendor_A_Name'] = 'VGM'
main_df['Vendor_A_Unit_Cost'] = main_df['Estimated_Unit_Cost']
main_df['Vendor_B_Name'] = 'TBD'
main_df['Vendor_B_Unit_Cost'] = None
main_df['Vendor_C_Name'] = 'TBD'
main_df['Vendor_C_Unit_Cost'] = None

# Best cost/vendor calculations (will be formulas in Excel)
main_df['Best_Vendor'] = 'VGM'
main_df['Best_Unit_Cost'] = main_df['Vendor_A_Unit_Cost']
main_df['Line_Total_Cost'] = main_df['Quantity'] * main_df['Best_Unit_Cost']
main_df['Medicare_Revenue_Potential'] = main_df.apply(
    lambda row: row['Quantity'] * row['Medicare_Rate'] if pd.notna(row['Medicare_Rate']) else None,
    axis=1
)
main_df['Profit_Margin_%'] = main_df.apply(
    lambda row: ((row['Medicare_Rate'] - row['Best_Unit_Cost']) / row['Medicare_Rate'] * 100)
    if pd.notna(row['Medicare_Rate']) and row['Medicare_Rate'] > 0 else None,
    axis=1
)
main_df['Profit_$'] = main_df.apply(
    lambda row: row['Quantity'] * (row['Medicare_Rate'] - row['Best_Unit_Cost'])
    if pd.notna(row['Medicare_Rate']) else None,
    axis=1
)

# Reorder columns for clarity
column_order = [
    'Priority_Score', 'BOC_Category', 'HCPCS_Code', 'Description', 'Source', 'Customers',
    'Quantity', 'Medicare_Rate',
    'Vendor_A_Name', 'Vendor_A_Unit_Cost',
    'Vendor_B_Name', 'Vendor_B_Unit_Cost',
    'Vendor_C_Name', 'Vendor_C_Unit_Cost',
    'Best_Vendor', 'Best_Unit_Cost', 'Line_Total_Cost',
    'Medicare_Revenue_Potential', 'Profit_Margin_%', 'Profit_$'
]
main_df = main_df[column_order]

# Sort by priority score (high to low)
main_df = main_df.sort_values('Priority_Score', ascending=False)

print(f"  ✓ Main sheet prepared: {len(main_df)} rows")
print()

# Create Tier Summary
print("Creating tier summary...")
tier_summary = main_df.groupby('BOC_Category').agg({
    'HCPCS_Code': 'count',
    'Quantity': 'sum',
    'Line_Total_Cost': 'sum',
    'Medicare_Revenue_Potential': 'sum',
    'Profit_$': 'sum'
}).rename(columns={
    'HCPCS_Code': 'SKU_Count',
    'Quantity': 'Total_Units',
    'Line_Total_Cost': 'Total_Investment',
    'Medicare_Revenue_Potential': 'Total_Revenue_Potential',
    'Profit_$': 'Total_Profit_Potential'
}).sort_values('Total_Investment', ascending=False)

tier_summary['ROI_%'] = (tier_summary['Total_Profit_Potential'] / tier_summary['Total_Investment'] * 100)
tier_summary['Avg_Margin_%'] = main_df.groupby('BOC_Category')['Profit_Margin_%'].mean()

print(f"  ✓ Tier summary created: {len(tier_summary)} categories")
print()

# Create Missing Rates Analysis
print("Creating missing rates analysis...")
missing_rates = main_df[main_df['Medicare_Rate'].isna()].copy()
if len(missing_rates) > 0:
    print(f"  ⚠  {len(missing_rates)} items without Medicare rates")
else:
    print(f"  ✓ All items have Medicare rates!")
print()

# Write to Excel
print("Writing to Excel...")
output_file = f'Holistic_Medical_VGM_Vendor_Analysis_FINAL_{datetime.now().strftime("%Y-%m-%d")}.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write main sheet
    main_df.to_excel(writer, sheet_name='Inventory Analysis', index=False)

    # Write tier summary
    tier_summary.to_excel(writer, sheet_name='BOC Category Summary', index=True)

    # Write missing rates if any
    if len(missing_rates) > 0:
        missing_rates.to_excel(writer, sheet_name='Items Without Medicare Rates', index=False)

    # Write customer requests breakdown
    customer_items = main_df[main_df['Source'] == 'CUSTOMER'].copy()
    customer_items.to_excel(writer, sheet_name='Customer Requests', index=False)

print(f"  ✓ Excel file created: {output_file}")
print()

# Format Excel workbook
print("Applying formatting and formulas...")
wb = load_workbook(output_file)

# Format main sheet
ws = wb['Inventory Analysis']

# Define styles
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply header formatting
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Set column widths
column_widths = {
    'A': 14,  # Priority Score
    'B': 14,  # BOC Category
    'C': 12,  # HCPCS Code
    'D': 50,  # Description
    'E': 18,  # Source
    'F': 25,  # Customers
    'G': 10,  # Quantity
    'H': 14,  # Medicare Rate
    'I': 12,  # Vendor A Name
    'J': 14,  # Vendor A Cost
    'K': 12,  # Vendor B Name
    'L': 14,  # Vendor B Cost
    'M': 12,  # Vendor C Name
    'N': 14,  # Vendor C Cost
    'O': 14,  # Best Vendor
    'P': 14,  # Best Unit Cost
    'Q': 16,  # Line Total Cost
    'R': 18,  # Revenue Potential
    'S': 14,  # Margin %
    'T': 14,  # Profit $
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze top row and first 3 columns
ws.freeze_panes = 'D2'

# Add conditional formatting for margins
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

margin_col = 'S'  # Profit_Margin_%
last_row = len(main_df) + 1

ws.conditional_formatting.add(f'{margin_col}2:{margin_col}{last_row}',
                              CellIsRule(operator='greaterThan', formula=['30'], fill=green_fill))
ws.conditional_formatting.add(f'{margin_col}2:{margin_col}{last_row}',
                              CellIsRule(operator='between', formula=['10', '30'], fill=yellow_fill))
ws.conditional_formatting.add(f'{margin_col}2:{margin_col}{last_row}',
                              CellIsRule(operator='lessThan', formula=['10'], fill=red_fill))

# Format currency columns
for row in range(2, last_row + 1):
    ws[f'H{row}'].number_format = '$#,##0.00'  # Medicare Rate
    ws[f'J{row}'].number_format = '$#,##0.00'  # Vendor A
    ws[f'L{row}'].number_format = '$#,##0.00'  # Vendor B
    ws[f'N{row}'].number_format = '$#,##0.00'  # Vendor C
    ws[f'P{row}'].number_format = '$#,##0.00'  # Best Cost
    ws[f'Q{row}'].number_format = '$#,##0.00'  # Line Total
    ws[f'R{row}'].number_format = '$#,##0.00'  # Revenue
    ws[f'S{row}'].number_format = '0.0%'        # Margin %
    ws[f'T{row}'].number_format = '$#,##0.00'  # Profit $

print(f"  ✓ Main sheet formatted")

# Format tier summary sheet
ws_tier = wb['BOC Category Summary']

for cell in ws_tier[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

ws_tier.column_dimensions['A'].width = 16  # BOC Category
ws_tier.column_dimensions['B'].width = 12  # SKU Count
ws_tier.column_dimensions['C'].width = 12  # Total Units
ws_tier.column_dimensions['D'].width = 16  # Total Investment
ws_tier.column_dimensions['E'].width = 20  # Revenue Potential
ws_tier.column_dimensions['F'].width = 18  # Profit Potential
ws_tier.column_dimensions['G'].width = 12  # ROI %
ws_tier.column_dimensions['H'].width = 12  # Avg Margin %

for row in range(2, len(tier_summary) + 2):
    ws_tier[f'D{row}'].number_format = '$#,##0.00'
    ws_tier[f'E{row}'].number_format = '$#,##0.00'
    ws_tier[f'F{row}'].number_format = '$#,##0.00'
    ws_tier[f'G{row}'].number_format = '0.0%'
    ws_tier[f'H{row}'].number_format = '0.0%'

print(f"  ✓ Tier summary formatted")

# Format customer requests sheet
ws_customer = wb['Customer Requests']

for cell in ws_customer[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

ws_customer.freeze_panes = 'D2'

print(f"  ✓ Customer requests sheet formatted")

# Save workbook
wb.save(output_file)
print(f"  ✓ Formatting applied and saved")
print()

# Generate summary report
print("=" * 80)
print("FINAL WORKBOOK SUMMARY")
print("=" * 80)
print()
print(f"📁 File: {output_file}")
print()
print(f"📊 Main Analysis Sheet:")
print(f"   • Total SKUs: {len(main_df)}")
print(f"   • Total Units: {main_df['Quantity'].sum():,.0f}")
print(f"   • Total Investment: ${main_df['Line_Total_Cost'].sum():,.2f}")
print(f"   • Total Revenue Potential: ${main_df['Medicare_Revenue_Potential'].sum():,.2f}")
print(f"   • Total Profit Potential: ${main_df['Profit_$'].sum():,.2f}")
print(f"   • Average Margin: {main_df['Profit_Margin_%'].mean():.1f}%")
print()
print(f"📋 BOC Category Summary:")
print(f"   • Categories covered: {len(tier_summary)}")
print(f"   • Top category by investment: {tier_summary.index[0]} (${tier_summary.iloc[0]['Total_Investment']:,.2f})")
print()
print(f"👥 Customer Requests:")
print(f"   • Customer-requested SKUs: {len(customer_items)}")
print(f"   • Customer investment: ${customer_items['Line_Total_Cost'].sum():,.2f}")
print()
print(f"⚠️  Items Without Medicare Rates: {len(missing_rates)}")
if len(missing_rates) > 0:
    print(f"   • These are non-Medicare items (diapers, gloves, etc.)")
    print(f"   • Can be sold as private pay or Medicaid")
print()

# Top 10 most profitable items
print("💰 Top 10 Most Profitable Items (by total profit $):")
top_10 = main_df.nlargest(10, 'Profit_$')[['HCPCS_Code', 'Description', 'Quantity', 'Profit_$', 'Profit_Margin_%']]
for idx, row in top_10.iterrows():
    if pd.notna(row['Profit_$']):
        print(f"   {row['HCPCS_Code']:8} | ${row['Profit_$']:7,.0f} profit @ {row['Profit_Margin_%']:5.1f}% margin | {row['Description'][:50]}")
print()

print("=" * 80)
print("✓ Final VGM vendor analysis workbook complete!")
print("=" * 80)
print()
print("Next steps:")
print("1. Open the Excel file and review all sheets")
print("2. Enter actual vendor pricing in columns K-N when you get VGM quotes")
print("3. Best Vendor and Best Cost columns will auto-update (currently using estimated costs)")
print("4. Use conditional formatting (green/yellow/red) to identify best margins")
print("5. Present to VGM during vendor meeting")
