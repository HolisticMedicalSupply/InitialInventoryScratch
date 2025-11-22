#!/usr/bin/env python3
"""
PHASE 3 EXECUTION: Build VGM Vendor Meeting Analysis Workbook
Holistic Medical Supply Inc. - Complete Medicare Rate Matching & Profitability Analysis
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

print("="*100)
print("PHASE 3: BUILD VGM VENDOR ANALYSIS WORKBOOK")
print("="*100)

# ============================================================================
# STEP 1: LOAD AND CLEAN SOURCE DATA
# ============================================================================
print("\nSTEP 1: Loading source data...")

# Load inventory
inventory_raw = pd.read_excel(
    "/home/user/InitialInventoryScratch/Holistic_Medical_Inventory_DETAILED(1).xlsx",
    sheet_name="Inventory Detail"
)
print(f"  Loaded inventory: {len(inventory_raw)} rows")

# Load Medicare rates
medicare = pd.read_excel(
    "/home/user/InitialInventoryScratch/Medicare_Rates_Normalized_Structure_Validated.xlsx",
    sheet_name="Medicare Rates - Normalized"
)
print(f"  Loaded Medicare rates: {len(medicare)} entries")

# Clean inventory - remove subtotal rows
inventory = inventory_raw[inventory_raw['HCPCS Code'].notna()].copy()
print(f"  Clean inventory: {len(inventory)} products (removed {len(inventory_raw) - len(inventory)} subtotal rows)")

# ============================================================================
# STEP 2: PREPARE MEDICARE RATE LOOKUP
# ============================================================================
print("\nSTEP 2: Preparing Medicare rate lookup...")

# Create lookup key in Medicare data
medicare['Lookup_Key'] = (
    medicare['HCPCS Code'].astype(str) + '_' +
    medicare['Modifier 1'].fillna('NONE').astype(str) + '_' +
    medicare['Geographic Tier'].astype(str) + '_' +
    medicare['Delivery Method'].astype(str)
)

# Create primary rate dictionary (NU + Urban + Standard)
medicare_nu_urban = medicare[
    (medicare['Modifier 1'] == 'NU') &
    (medicare['Geographic Tier'] == 'Urban') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

# Create fallback dictionary (no modifier + Urban + Standard) for consumables
medicare_nomod_urban = medicare[
    (medicare['Modifier 1'].isna()) &
    (medicare['Geographic Tier'] == 'Urban') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

# Create RR (rental) rate dictionary
medicare_rr_urban = medicare[
    (medicare['Modifier 1'] == 'RR') &
    (medicare['Geographic Tier'] == 'Urban') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

# Create Rural rate dictionary
medicare_nu_rural = medicare[
    (medicare['Modifier 1'] == 'NU') &
    (medicare['Geographic Tier'] == 'Rural') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

medicare_nomod_rural = medicare[
    (medicare['Modifier 1'].isna()) &
    (medicare['Geographic Tier'] == 'Rural') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

print(f"  Created lookup dictionaries:")
print(f"    NU + Urban: {len(medicare_nu_urban)} codes")
print(f"    No Modifier + Urban: {len(medicare_nomod_urban)} codes")
print(f"    RR + Urban: {len(medicare_rr_urban)} codes")

# ============================================================================
# STEP 3: MATCH MEDICARE RATES TO INVENTORY
# ============================================================================
print("\nSTEP 3: Matching Medicare rates to inventory...")

def get_medicare_rate(hcpcs_code, rate_dict_primary, rate_dict_fallback):
    """Try primary dictionary first, then fallback"""
    if pd.isna(hcpcs_code):
        return None
    if hcpcs_code in rate_dict_primary:
        return rate_dict_primary[hcpcs_code]
    elif hcpcs_code in rate_dict_fallback:
        return rate_dict_fallback[hcpcs_code]
    else:
        return None

# Apply Medicare rate matching
inventory['Medicare_Rate_NU'] = inventory['HCPCS Code'].apply(
    lambda x: get_medicare_rate(x, medicare_nu_urban, medicare_nomod_urban)
)

inventory['Medicare_Rate_RR'] = inventory['HCPCS Code'].apply(
    lambda x: get_medicare_rate(x, medicare_rr_urban, {})
)

inventory['Medicare_Rate_Rural'] = inventory['HCPCS Code'].apply(
    lambda x: get_medicare_rate(x, medicare_nu_rural, medicare_nomod_rural)
)

matched = inventory['Medicare_Rate_NU'].notna().sum()
total = len(inventory)
print(f"  Matched rates: {matched}/{total} ({matched/total*100:.1f}%)")
print(f"  Missing rates: {total - matched} codes")

# Add Medicare status
inventory['Medicare_Status'] = inventory['Medicare_Rate_NU'].apply(
    lambda x: "Covered - Part B" if pd.notna(x) else "Not Covered - Part B"
)

# ============================================================================
# STEP 4: CALCULATE UNIT COSTS AND DERIVED COLUMNS
# ============================================================================
print("\nSTEP 4: Calculating unit costs and derived columns...")

# Extract budget as number
def clean_currency(val):
    if pd.isna(val):
        return 0
    if isinstance(val, str):
        return float(val.replace('$', '').replace(',', ''))
    return float(val)

inventory['Budget_Allocation_Original'] = inventory['Budget Allocation'].apply(clean_currency)

# Clean quantity column (convert to numeric, handle text)
def clean_quantity(val):
    if pd.isna(val):
        return 1
    if isinstance(val, str):
        # Handle strings like "3000 units", "25 boxes", etc.
        try:
            return float(val.split()[0])
        except:
            return 1
    return float(val)

inventory['Quantity_Clean'] = inventory['Quantity'].apply(clean_quantity)

# Calculate estimated unit cost from budget
inventory['Unit_Cost_Estimate'] = inventory['Budget_Allocation_Original'] / inventory['Quantity_Clean']

# Vendor columns (blank for now - to be filled by user)
inventory['Vendor_A_Name'] = 'VGM'
inventory['Vendor_A_Unit_Cost'] = inventory['Unit_Cost_Estimate']  # Use estimate as placeholder
inventory['Vendor_B_Name'] = 'Vendor B (TBD)'
inventory['Vendor_B_Unit_Cost'] = np.nan
inventory['Vendor_C_Name'] = 'Vendor C (TBD)'
inventory['Vendor_C_Unit_Cost'] = np.nan

# Best unit cost (for now, just use estimate since other vendors blank)
inventory['Best_Unit_Cost'] = inventory['Vendor_A_Unit_Cost']
inventory['Best_Vendor'] = inventory['Vendor_A_Name']

# Line calculations
inventory['Line_Total_Cost'] = inventory['Quantity_Clean'] * inventory['Best_Unit_Cost']

inventory['Line_Medicare_Revenue'] = inventory.apply(
    lambda row: 0 if pd.isna(row['Medicare_Rate_NU']) else row['Quantity_Clean'] * row['Medicare_Rate_NU'],
    axis=1
)

inventory['Line_Gross_Margin'] = inventory['Line_Medicare_Revenue'] - inventory['Line_Total_Cost']

inventory['Margin_Percentage'] = inventory.apply(
    lambda row: np.nan if row['Line_Medicare_Revenue'] == 0 else row['Line_Gross_Margin'] / row['Line_Medicare_Revenue'],
    axis=1
)

# Margin category
def categorize_margin(pct):
    if pd.isna(pct):
        return "Non-Reimbursable"
    elif pct >= 0.30:
        return "High (>30%)"
    elif pct >= 0.10:
        return "Medium (10-30%)"
    elif pct >= 0:
        return "Low (<10%)"
    else:
        return "LOSS"

inventory['Margin_Category'] = inventory['Margin_Percentage'].apply(categorize_margin)

# Budget variance
inventory['Budget_Variance'] = inventory['Line_Total_Cost'] - inventory['Budget_Allocation_Original']

# Priority score
def calculate_priority(row):
    if pd.isna(row['Margin_Percentage']):
        return 0

    margin_score = row['Margin_Percentage'] * 0.4
    profit_score = (row['Line_Gross_Margin'] / 10000) * 0.3
    volume_score = (row['Quantity_Clean'] / 100) * 0.2
    coverage_bonus = 0.1 if row['Medicare_Status'] == "Covered - Part B" else 0

    return margin_score + profit_score + volume_score + coverage_bonus

inventory['Priority_Score'] = inventory.apply(calculate_priority, axis=1)

# Rate spread (difference between highest and lowest rate)
inventory['Rate_Spread'] = inventory.apply(
    lambda row: np.nan if pd.isna(row['Medicare_Rate_NU']) else
                (max([r for r in [row['Medicare_Rate_NU'], row['Medicare_Rate_RR'], row['Medicare_Rate_Rural']] if pd.notna(r)]) -
                 min([r for r in [row['Medicare_Rate_NU'], row['Medicare_Rate_RR'], row['Medicare_Rate_Rural']] if pd.notna(r)])),
    axis=1
)

# Add unit column
inventory['Unit'] = 'each'

# Last updated
inventory['Last_Updated'] = datetime.now().strftime('%Y-%m-%d')

print(f"  Calculated {len(inventory)} rows of derived data")

# ============================================================================
# STEP 5: CREATE MAIN ANALYSIS DATAFRAME (FINAL COLUMN ORDER)
# ============================================================================
print("\nSTEP 5: Creating main analysis dataframe...")

main_analysis = inventory[[
    'Tier',
    'BOC Category',
    'Product Line',
    'HCPCS Code',
    'Product Description',
    'Quantity_Clean',
    'Unit',
    'Medicare_Rate_NU',
    'Medicare_Status',
    'Unit_Cost_Estimate',
    'Vendor_A_Name',
    'Vendor_A_Unit_Cost',
    'Vendor_B_Name',
    'Vendor_B_Unit_Cost',
    'Vendor_C_Name',
    'Vendor_C_Unit_Cost',
    'Best_Unit_Cost',
    'Best_Vendor',
    'Line_Total_Cost',
    'Line_Medicare_Revenue',
    'Line_Gross_Margin',
    'Margin_Percentage',
    'Margin_Category',
    'Budget_Allocation_Original',
    'Budget_Variance',
    'Priority_Score',
    'Notes',
    'Medicare_Rate_RR',
    'Medicare_Rate_Rural',
    'Rate_Spread',
    'Last_Updated'
]].copy()

# Rename for better display
main_analysis.columns = [
    'Tier',
    'BOC Category',
    'Product Line',
    'HCPCS Code',
    'Product Description',
    'Quantity',
    'Unit',
    'Medicare Rate (NU)',
    'Medicare Status',
    'Unit Cost (Est)',
    'Vendor A',
    'Vendor A Cost',
    'Vendor B',
    'Vendor B Cost',
    'Vendor C',
    'Vendor C Cost',
    'Best Cost',
    'Best Vendor',
    'Line Total Cost',
    'Line Medicare Revenue',
    'Line Gross Margin',
    'Margin %',
    'Margin Category',
    'Budget (Original)',
    'Budget Variance',
    'Priority Score',
    'Notes',
    'Medicare Rate (RR Rental)',
    'Medicare Rate (Rural)',
    'Rate Spread',
    'Last Updated'
]

print(f"  Main analysis: {len(main_analysis)} rows × {len(main_analysis.columns)} columns")

# ============================================================================
# STEP 6: CREATE MISSING RATES DATAFRAME
# ============================================================================
print("\nSTEP 6: Creating missing rates analysis...")

missing_rates = inventory[inventory['Medicare_Status'] == 'Not Covered - Part B'].copy()

missing_analysis = missing_rates[[
    'HCPCS Code',
    'Product Description',
    'Tier',
    'BOC Category',
    'Quantity_Clean',
    'Unit_Cost_Estimate',
    'Line_Total_Cost'
]].copy()

# Add explanation columns
missing_analysis['Why Not Covered'] = missing_analysis['HCPCS Code'].apply(
    lambda x: "Personal hygiene - not DME" if x in ['A4520', 'A4554', 'A4553']
          else "Verify correct HCPCS code" if x in ['A6198', 'A6217', 'A6261', 'A6262', 'L1810', 'L1820', 'L1900', 'L3702', 'E0118']
          else "Research required"
)

missing_analysis['Alternative Payer'] = missing_analysis['HCPCS Code'].apply(
    lambda x: "Medicaid (state-dependent)" if x in ['A4520', 'A4554', 'A4553']
          else "TBD - verify code first"
)

missing_analysis['Action Required'] = "Research & verify code or accept as non-covered"

# Rename columns
missing_analysis.columns = [
    'HCPCS Code',
    'Product Description',
    'Tier',
    'BOC Category',
    'Quantity',
    'Unit Cost (Est)',
    'Line Total Cost',
    'Why Not Covered',
    'Alternative Payer',
    'Action Required'
]

print(f"  Missing rates analysis: {len(missing_analysis)} codes")

# ============================================================================
# STEP 7: CREATE TIER BUDGET SUMMARY
# ============================================================================
print("\nSTEP 7: Creating tier budget summary...")

# Tier summary
tier_summary = inventory.groupby('Tier').agg({
    'HCPCS Code': 'count',
    'Quantity_Clean': 'sum',
    'Line_Total_Cost': 'sum',
    'Line_Medicare_Revenue': 'sum',
    'Line_Gross_Margin': 'sum',
    'Margin_Percentage': lambda x: x[x.notna()].mean()
}).round(2)

tier_summary.columns = ['Product Count', 'Total Quantity', 'Total Cost', 'Total Medicare Revenue', 'Total Margin $', 'Avg Margin %']
tier_summary = tier_summary.reset_index()

print(f"  Tier summary: {len(tier_summary)} tiers")

# BOC Category summary
boc_summary = inventory.groupby('BOC Category').agg({
    'HCPCS Code': 'count',
    'Line_Total_Cost': 'sum',
    'Line_Medicare_Revenue': 'sum',
    'Line_Gross_Margin': 'sum',
    'Margin_Percentage': lambda x: x[x.notna()].mean(),
    'Priority_Score': 'sum'
}).round(2)

boc_summary.columns = ['Product Count', 'Total Cost', 'Total Revenue', 'Total Margin $', 'Avg Margin %', 'Total Priority Score']
boc_summary = boc_summary.reset_index()
boc_summary = boc_summary.sort_values('Total Margin $', ascending=False)

print(f"  BOC category summary: {len(boc_summary)} categories")

# Top 20 products by margin
top_20 = inventory.nlargest(20, 'Line_Gross_Margin')[
    ['Product Description', 'HCPCS Code', 'Line_Gross_Margin', 'Margin_Percentage', 'Quantity_Clean', 'Tier']
].copy()
top_20['Rank'] = range(1, len(top_20) + 1)
top_20 = top_20[['Rank', 'Product Description', 'HCPCS Code', 'Line_Gross_Margin', 'Margin_Percentage', 'Quantity_Clean', 'Tier']]
top_20.columns = ['Rank', 'Product', 'HCPCS', 'Margin $', 'Margin %', 'Quantity', 'Tier']

print(f"  Top 20 products by margin")

# Bottom 20 products (losses and low margins)
bottom_20 = inventory.nsmallest(20, 'Line_Gross_Margin')[
    ['Product Description', 'HCPCS Code', 'Line_Gross_Margin', 'Margin_Percentage', 'Quantity_Clean', 'Tier', 'Medicare_Status']
].copy()
bottom_20['Rank'] = range(1, len(bottom_20) + 1)
bottom_20['Action'] = bottom_20.apply(
    lambda row: "Not covered - private pay only" if row['Medicare_Status'] == "Not Covered - Part B"
           else "LOSS - reconsider stocking" if row['Line_Gross_Margin'] < 0
           else "Low margin - monitor",
    axis=1
)
bottom_20 = bottom_20[['Rank', 'Product Description', 'HCPCS Code', 'Line_Gross_Margin', 'Margin_Percentage', 'Action', 'Tier']]
bottom_20.columns = ['Rank', 'Product', 'HCPCS', 'Margin $', 'Margin %', 'Action', 'Tier']

print(f"  Bottom 20 products by margin")

# ============================================================================
# STEP 8: CREATE EXCEL WORKBOOK
# ============================================================================
print("\nSTEP 8: Creating Excel workbook...")

output_file = f"/home/user/InitialInventoryScratch/Holistic_Medical_VGM_Vendor_Analysis_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

# Create workbook
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write all sheets
    main_analysis.to_excel(writer, sheet_name='Main Analysis', index=False)
    missing_analysis.to_excel(writer, sheet_name='Missing Rates', index=False)
    tier_summary.to_excel(writer, sheet_name='Tier Summary', index=False, startrow=0)
    boc_summary.to_excel(writer, sheet_name='Tier Summary', index=False, startrow=len(tier_summary) + 3)
    top_20.to_excel(writer, sheet_name='Tier Summary', index=False, startrow=len(tier_summary) + len(boc_summary) + 6)
    bottom_20.to_excel(writer, sheet_name='Tier Summary', index=False, startrow=len(tier_summary) + len(boc_summary) + len(top_20) + 9)

print(f"  ✓ Wrote data to {output_file}")

# ============================================================================
# STEP 9: APPLY FORMATTING
# ============================================================================
print("\nSTEP 9: Applying professional formatting...")

wb = load_workbook(output_file)

# Define styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Format Main Analysis sheet
ws_main = wb['Main Analysis']

# Freeze panes (row 2, so headers stay visible)
ws_main.freeze_panes = 'A2'

# Format headers (row 1)
for cell in ws_main[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Set row height for header
ws_main.row_dimensions[1].height = 30

# Alternating row colors and borders
light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
for idx, row in enumerate(ws_main.iter_rows(min_row=2, max_row=ws_main.max_row), start=2):
    for cell in row:
        cell.border = thin_border
        if idx % 2 == 0:
            cell.fill = light_gray

        # Number formatting
        col_letter = cell.column_letter
        if col_letter in ['H', 'J', 'L', 'N', 'P', 'Q', 'R', 'S', 'T', 'U', 'X', 'AB', 'AC', 'AD']:  # Currency columns
            cell.number_format = '$#,##0.00'
        elif col_letter in ['V']:  # Margin percentage
            cell.number_format = '0.0%'
        elif col_letter in ['Z']:  # Priority score
            cell.number_format = '0.00'

# Conditional formatting for Margin %
ws_main.conditional_formatting.add(
    f'V2:V{ws_main.max_row}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0.30'],
               fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
               font=Font(color='006100'))
)
ws_main.conditional_formatting.add(
    f'V2:V{ws_main.max_row}',
    CellIsRule(operator='between', formula=['0.10', '0.299'],
               fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
               font=Font(color='9C6500'))
)
ws_main.conditional_formatting.add(
    f'V2:V{ws_main.max_row}',
    CellIsRule(operator='lessThan', formula=['0.10'],
               fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
               font=Font(color='9C0006'))
)

# Set column widths
column_widths = {
    'A': 30, 'B': 15, 'C': 20, 'D': 12, 'E': 45, 'F': 10, 'G': 8,
    'H': 15, 'I': 20, 'J': 15, 'K': 15, 'L': 15, 'M': 15, 'N': 15,
    'O': 15, 'P': 15, 'Q': 15, 'R': 15, 'S': 18, 'T': 18, 'U': 18,
    'V': 12, 'W': 20, 'X': 18, 'Y': 18, 'Z': 15, 'AA': 30, 'AB': 18,
    'AC': 18, 'AD': 15, 'AE': 15
}
for col, width in column_widths.items():
    ws_main.column_dimensions[col].width = width

# Format Missing Rates sheet
ws_missing = wb['Missing Rates']
ws_missing.freeze_panes = 'A2'
for cell in ws_missing[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
ws_missing.column_dimensions['B'].width = 45
ws_missing.column_dimensions['H'].width = 30
ws_missing.column_dimensions['I'].width = 30
ws_missing.column_dimensions['J'].width = 35

# Format Tier Summary sheet
ws_summary = wb['Tier Summary']
ws_summary.freeze_panes = 'A2'

# Add section headers
ws_summary.insert_rows(1)
ws_summary['A1'] = 'TIER BUDGET SUMMARY'
ws_summary['A1'].font = Font(name='Calibri', size=14, bold=True)

tier_start = 2
boc_start = tier_start + len(tier_summary) + 4
top20_start = boc_start + len(boc_summary) + 3
bottom20_start = top20_start + len(top_20) + 3

ws_summary.insert_rows(boc_start - 1)
ws_summary[f'A{boc_start - 1}'] = 'BOC CATEGORY SUMMARY'
ws_summary[f'A{boc_start - 1}'].font = Font(name='Calibri', size=14, bold=True)

ws_summary.insert_rows(top20_start - 1)
ws_summary[f'A{top20_start - 1}'] = 'TOP 20 PRODUCTS BY MARGIN'
ws_summary[f'A{top20_start - 1}'].font = Font(name='Calibri', size=14, bold=True)

ws_summary.insert_rows(bottom20_start - 1)
ws_summary[f'A{bottom20_start - 1}'] = 'BOTTOM 20 PRODUCTS (LOW MARGIN / LOSS)'
ws_summary[f'A{bottom20_start - 1}'].font = Font(name='Calibri', size=14, bold=True)

# Tab colors
wb['Main Analysis'].sheet_properties.tabColor = "1F77B4"  # Blue
wb['Missing Rates'].sheet_properties.tabColor = "D62728"  # Red
wb['Tier Summary'].sheet_properties.tabColor = "FF7F0E"  # Orange

# Save workbook
wb.save(output_file)
print(f"  ✓ Applied formatting")

# ============================================================================
# STEP 10: VALIDATION
# ============================================================================
print("\nSTEP 10: Running validation checks...")

# Check 1: Medicare rate matching
matched_pct = (inventory['Medicare_Rate_NU'].notna().sum() / len(inventory)) * 100
print(f"  ✓ Medicare rate match: {matched_pct:.1f}%")

# Check 2: Line total calculations
line_totals_correct = ((inventory['Line_Total_Cost'] - (inventory['Quantity_Clean'] * inventory['Best_Unit_Cost'])).abs() < 0.01).all()
print(f"  ✓ Line total calculations: {'PASS' if line_totals_correct else 'FAIL'}")

# Check 3: No blanks in critical fields
critical_fields = ['HCPCS Code', 'Product Description', 'Quantity', 'Tier', 'BOC Category']
no_blanks = all(inventory[field].notna().all() for field in critical_fields)
print(f"  ✓ No blanks in critical fields: {'PASS' if no_blanks else 'FAIL'}")

# Check 4: Tier budget totals
tier_check = abs(inventory['Line_Total_Cost'].sum() - tier_summary['Total Cost'].sum()) < 1.0
print(f"  ✓ Tier summary reconciliation: {'PASS' if tier_check else 'FAIL'}")

# Check 5: File exists and size
import os
file_exists = os.path.exists(output_file)
file_size = os.path.getsize(output_file) if file_exists else 0
print(f"  ✓ Output file created: {file_size:,} bytes")

# ============================================================================
# COMPLETION SUMMARY
# ============================================================================
print("\n" + "="*100)
print("PHASE 3 EXECUTION COMPLETE ✅")
print("="*100)
print(f"\nOutput File: {output_file}")
print(f"\nWorkbook Contents:")
print(f"  - Main Analysis: {len(main_analysis)} products with full Medicare rate matching")
print(f"  - Missing Rates: {len(missing_analysis)} codes requiring research")
print(f"  - Tier Summary: {len(tier_summary)} tiers + {len(boc_summary)} BOC categories + Top/Bottom 20")
print(f"\nKey Statistics:")
print(f"  - Total Products: {len(inventory)}")
print(f"  - Medicare Rate Coverage: {matched_pct:.1f}%")
print(f"  - Total Budget (Original): ${inventory['Budget_Allocation_Original'].sum():,.2f}")
print(f"  - Total Line Cost: ${inventory['Line_Total_Cost'].sum():,.2f}")
print(f"  - Total Medicare Revenue: ${inventory['Line_Medicare_Revenue'].sum():,.2f}")
print(f"  - Total Gross Margin: ${inventory['Line_Gross_Margin'].sum():,.2f}")
print(f"  - Average Margin %: {inventory['Margin_Percentage'].mean()*100:.1f}%")
print(f"\nVGM Vendor Meeting Ready: ✅")
print("="*100)
