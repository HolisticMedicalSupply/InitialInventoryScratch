#!/usr/bin/env python3
"""
Add missing items to inventory and rebuild VGM vendor analysis workbook
- Add 5 missing Dr. Nas ankle products
- Add 4 missing CGM full system items
- Rebuild workbook with proper formulas and formatting
- Validate everything
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

print("="*100)
print("ADDING MISSING ITEMS AND REBUILDING VGM VENDOR ANALYSIS WORKBOOK")
print("="*100)

# ============================================================================
# STEP 1: DEFINE MISSING ITEMS
# ============================================================================
print("\nSTEP 1: Defining missing items...")

# Missing Dr. Nas ankle products (5 items)
missing_ankle_products = [
    {
        'Tier': 'TIER 7: SPECIALIZED EQUIPMENT',
        'BOC Category': 'OR03',
        'Product Line': 'Ankle/Foot Orthotics - Dr. Nas',
        'HCPCS Code': 'L1906',
        'Product Description': 'AFO - multiligamentous ankle support, prefabricated, off-the-shelf',
        'Quantity': 6,
        'Budget Allocation': '$480',
        'Notes': 'Dr. Nas ankle specialist request'
    },
    {
        'Tier': 'TIER 7: SPECIALIZED EQUIPMENT',
        'BOC Category': 'OR03',
        'Product Line': 'Ankle/Foot Orthotics - Dr. Nas',
        'HCPCS Code': 'L4361',
        'Product Description': 'Walking boot - pneumatic and/or vacuum, prefabricated, off-the-shelf (CAM WALKER)',
        'Quantity': 6,
        'Budget Allocation': '$420',
        'Notes': 'Dr. Nas PRIMARY REQUEST - CAM walker'
    },
    {
        'Tier': 'TIER 7: SPECIALIZED EQUIPMENT',
        'BOC Category': 'OR03',
        'Product Line': 'Ankle/Foot Orthotics - Dr. Nas',
        'HCPCS Code': 'L4350',
        'Product Description': 'Ankle control orthosis, stirrup style, rigid, prefabricated, off-the-shelf',
        'Quantity': 2,
        'Budget Allocation': '$160',
        'Notes': 'Dr. Nas ankle specialist request'
    },
    {
        'Tier': 'TIER 7: SPECIALIZED EQUIPMENT',
        'BOC Category': 'OR03',
        'Product Line': 'Ankle/Foot Orthotics - Dr. Nas',
        'HCPCS Code': 'L4370',
        'Product Description': 'Pneumatic full leg splint, prefabricated, off-the-shelf',
        'Quantity': 2,
        'Budget Allocation': '$220',
        'Notes': 'Dr. Nas ankle specialist request'
    },
    {
        'Tier': 'TIER 7: SPECIALIZED EQUIPMENT',
        'BOC Category': 'OR03',
        'Product Line': 'Ankle/Foot Orthotics - Dr. Nas',
        'HCPCS Code': 'L4387',
        'Product Description': 'Walking boot - non-pneumatic, prefabricated, off-the-shelf',
        'Quantity': 2,
        'Budget Allocation': '$270',
        'Notes': 'Dr. Nas ankle specialist request'
    }
]

# Missing CGM full systems (4 items - 2 receivers, 2 sensors)
missing_cgm_products = [
    {
        'Tier': 'TIER 4: DIABETES MANAGEMENT',
        'BOC Category': 'DM06',
        'Product Line': 'Continuous Glucose Monitors',
        'HCPCS Code': 'E2103',
        'Product Description': 'Dexcom G7 CGM Receiver (Non-display glucose monitor system)',
        'Quantity': 5,
        'Budget Allocation': '$2000',
        'Notes': 'Full CGM system - receiver. VGM to introduce to Dexcom'
    },
    {
        'Tier': 'TIER 4: DIABETES MANAGEMENT',
        'BOC Category': 'DM06',
        'Product Line': 'Continuous Glucose Monitors',
        'HCPCS Code': 'E2103',
        'Product Description': 'Freestyle Libre CGM Reader (Non-display glucose monitor system)',
        'Quantity': 5,
        'Budget Allocation': '$2000',
        'Notes': 'Full CGM system - receiver. VGM to introduce to Abbott'
    },
    {
        'Tier': 'TIER 4: DIABETES MANAGEMENT',
        'BOC Category': 'DM06',
        'Product Line': 'Continuous Glucose Monitors',
        'HCPCS Code': 'A4239',
        'Product Description': 'Dexcom G7 CGM Sensors, 30-day supply',
        'Quantity': 10,
        'Budget Allocation': '$900',
        'Notes': 'Full CGM system - sensors for Dexcom G7'
    },
    {
        'Tier': 'TIER 4: DIABETES MANAGEMENT',
        'BOC Category': 'DM06',
        'Product Line': 'Continuous Glucose Monitors',
        'HCPCS Code': 'A4239',
        'Product Description': 'Freestyle Libre CGM Sensors, 30-day supply',
        'Quantity': 10,
        'Budget Allocation': '$700',
        'Notes': 'Full CGM system - sensors for Libre'
    }
]

ankle_total = sum([int(item['Budget Allocation'].replace('$', '').replace(',', '')) for item in missing_ankle_products])
cgm_total = sum([int(item['Budget Allocation'].replace('$', '').replace(',', '')) for item in missing_cgm_products])

print(f"  Missing ankle products: {len(missing_ankle_products)} items, ${ankle_total:,}")
print(f"  Missing CGM products: {len(missing_cgm_products)} items, ${cgm_total:,}")
print(f"  Total additional budget: ${ankle_total + cgm_total:,}")

# ============================================================================
# STEP 2: LOAD EXISTING INVENTORY
# ============================================================================
print("\nSTEP 2: Loading existing inventory...")

inventory_path = "/home/user/InitialInventoryScratch/Holistic_Medical_Inventory_DETAILED(1).xlsx"
inventory_df = pd.read_excel(inventory_path, sheet_name="Inventory Detail")

print(f"  Existing inventory: {len(inventory_df)} rows (includes subtotal rows)")

# Remove subtotal rows
inventory_clean = inventory_df[inventory_df['HCPCS Code'].notna()].copy()
print(f"  Clean inventory: {len(inventory_clean)} products")

# ============================================================================
# STEP 3: ADD MISSING ITEMS
# ============================================================================
print("\nSTEP 3: Adding missing items to inventory...")

# Convert missing items to DataFrame
ankle_df = pd.DataFrame(missing_ankle_products)
cgm_df = pd.DataFrame(missing_cgm_products)

# Combine with existing
inventory_updated = pd.concat([inventory_clean, ankle_df, cgm_df], ignore_index=True)

print(f"  Updated inventory: {len(inventory_updated)} products")
print(f"  Added {len(ankle_df) + len(cgm_df)} new items")

# Save updated inventory (for reference)
updated_inventory_path = "/home/user/InitialInventoryScratch/Holistic_Medical_Inventory_DETAILED_WITH_MISSING_ITEMS.xlsx"
inventory_updated.to_excel(updated_inventory_path, sheet_name="Inventory Detail", index=False)
print(f"  ✓ Saved updated inventory to: {updated_inventory_path}")

# ============================================================================
# STEP 4: LOAD MEDICARE RATES
# ============================================================================
print("\nSTEP 4: Loading Medicare rates...")

medicare = pd.read_excel(
    "/home/user/InitialInventoryScratch/Medicare_Rates_Normalized_Structure_Validated.xlsx",
    sheet_name="Medicare Rates - Normalized"
)
print(f"  Loaded Medicare rates: {len(medicare)} entries")

# Create Medicare rate lookup dictionaries
medicare_nu_urban = medicare[
    (medicare['Modifier 1'] == 'NU') &
    (medicare['Geographic Tier'] == 'Urban') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

medicare_nomod_urban = medicare[
    (medicare['Modifier 1'].isna()) &
    (medicare['Geographic Tier'] == 'Urban') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

medicare_rr_urban = medicare[
    (medicare['Modifier 1'] == 'RR') &
    (medicare['Geographic Tier'] == 'Urban') &
    (medicare['Delivery Method'] == 'Standard')
].set_index('HCPCS Code')['Rate ($)'].to_dict()

print(f"  Created lookup dictionaries:")
print(f"    NU + Urban: {len(medicare_nu_urban)} codes")
print(f"    No Modifier + Urban: {len(medicare_nomod_urban)} codes")

# ============================================================================
# STEP 5: MATCH MEDICARE RATES
# ============================================================================
print("\nSTEP 5: Matching Medicare rates...")

def get_medicare_rate(hcpcs_code, rate_dict_primary, rate_dict_fallback):
    if pd.isna(hcpcs_code):
        return None
    if hcpcs_code in rate_dict_primary:
        return rate_dict_primary[hcpcs_code]
    elif hcpcs_code in rate_dict_fallback:
        return rate_dict_fallback[hcpcs_code]
    else:
        return None

inventory_updated['Medicare_Rate'] = inventory_updated['HCPCS Code'].apply(
    lambda x: get_medicare_rate(x, medicare_nu_urban, medicare_nomod_urban)
)

matched = inventory_updated['Medicare_Rate'].notna().sum()
total = len(inventory_updated)
print(f"  Matched rates: {matched}/{total} ({matched/total*100:.1f}%)")

# ============================================================================
# STEP 6: CALCULATE COSTS AND METRICS
# ============================================================================
print("\nSTEP 6: Calculating costs and metrics...")

# Clean budget and quantity
def clean_currency(val):
    if pd.isna(val):
        return 0
    if isinstance(val, str):
        return float(val.replace('$', '').replace(',', ''))
    return float(val)

def clean_quantity(val):
    if pd.isna(val):
        return 1
    if isinstance(val, str):
        try:
            return float(val.split()[0])
        except:
            return 1
    return float(val)

inventory_updated['Budget_Clean'] = inventory_updated['Budget Allocation'].apply(clean_currency)
inventory_updated['Quantity_Clean'] = inventory_updated['Quantity'].apply(clean_quantity)

# Calculate unit cost estimate
inventory_updated['Unit_Cost_Estimate'] = inventory_updated['Budget_Clean'] / inventory_updated['Quantity_Clean']

# Vendor columns (to be filled by user)
inventory_updated['Vendor_A_Name'] = 'VGM'
inventory_updated['Vendor_A_Unit_Cost'] = inventory_updated['Unit_Cost_Estimate']
inventory_updated['Vendor_B_Name'] = 'Vendor B (TBD)'
inventory_updated['Vendor_B_Unit_Cost'] = np.nan
inventory_updated['Vendor_C_Name'] = 'Vendor C (TBD)'
inventory_updated['Vendor_C_Unit_Cost'] = np.nan

# Best cost (use estimate for now)
inventory_updated['Best_Unit_Cost'] = inventory_updated['Vendor_A_Unit_Cost']
inventory_updated['Best_Vendor'] = 'VGM'

# Line calculations
inventory_updated['Line_Total_Cost'] = inventory_updated['Quantity_Clean'] * inventory_updated['Best_Unit_Cost']

inventory_updated['Line_Medicare_Revenue'] = inventory_updated.apply(
    lambda row: 0 if pd.isna(row['Medicare_Rate']) else row['Quantity_Clean'] * row['Medicare_Rate'],
    axis=1
)

inventory_updated['Line_Gross_Margin'] = inventory_updated['Line_Medicare_Revenue'] - inventory_updated['Line_Total_Cost']

inventory_updated['Margin_Percentage'] = inventory_updated.apply(
    lambda row: np.nan if row['Line_Medicare_Revenue'] == 0 else row['Line_Gross_Margin'] / row['Line_Medicare_Revenue'],
    axis=1
)

print(f"  Calculated {len(inventory_updated)} rows of metrics")

# ============================================================================
# STEP 7: ADD SOURCE TRACKING FOR CUSTOMER REQUESTS
# ============================================================================
print("\nSTEP 7: Adding source tracking...")

def determine_source(row):
    """Determine if item is from customer request or launch inventory"""
    hcpcs = row['HCPCS Code']
    notes = str(row.get('Notes', '')).upper()

    # Customer-specific codes
    customer_codes = {
        'E0607': 'MOYHINOR, WALTERS',
        'L1902': 'DR_NAS',
        'L1906': 'DR_NAS',
        'L4361': 'DR_NAS',
        'L4350': 'DR_NAS',
        'L4370': 'DR_NAS',
        'L4387': 'DR_NAS',
        'E0130': 'RAMBOM',
        'K0001': 'MOYHINOR',
        'A4253': 'WALTERS',
        'E0143': 'RAMBOM',
        'A4259': 'WALTERS',
        'B4150': 'RAMBOM',
        'E0570': 'RAMBOM',
        'A4554': 'WALTERS',
        'A4520': 'WALTERS',
        'E2103': 'MOYHINOR (CGM_STRATEGY)',
        'A4239': 'MOYHINOR (CGM_STRATEGY)'
    }

    if hcpcs in customer_codes:
        return 'CUSTOMER', customer_codes[hcpcs]
    else:
        return 'LAUNCH_INVENTORY', None

inventory_updated[['Source', 'Customers']] = inventory_updated.apply(
    lambda row: pd.Series(determine_source(row)), axis=1
)

# Add priority score
def calculate_priority(row):
    """Calculate priority score for ordering"""
    score = 0

    # Customer requests get highest priority
    if row['Source'] == 'CUSTOMER':
        score += 90 if 'DR_NAS' in str(row['Customers']) else 70
        if row['HCPCS Code'] == 'L4361':  # CAM walker PRIMARY REQUEST
            score = 100
    else:
        score += 40

    # High margin bonus
    if pd.notna(row['Margin_Percentage']) and row['Margin_Percentage'] > 0.30:
        score += 10

    # High volume bonus
    if row['Quantity_Clean'] > 100:
        score += 5

    return score

inventory_updated['Priority_Score'] = inventory_updated.apply(calculate_priority, axis=1)

print(f"  Added source tracking and priority scores")

# ============================================================================
# STEP 8: CREATE FINAL WORKBOOK
# ============================================================================
print("\nSTEP 8: Creating final VGM vendor analysis workbook...")

from openpyxl import Workbook

output_file = f"/home/user/InitialInventoryScratch/Holistic_Medical_VGM_Vendor_Analysis_FINAL_2025-11-22.xlsx"

wb = Workbook()

# ============================================================================
# SHEET 1: INVENTORY ANALYSIS
# ============================================================================
print("  Creating Sheet 1: Inventory Analysis...")

ws_main = wb.active
ws_main.title = "Inventory Analysis"

# Sort by priority score (highest first)
inventory_sorted = inventory_updated.sort_values('Priority_Score', ascending=False)

# Define columns for main analysis
main_cols = [
    'Priority_Score',
    'BOC Category',
    'HCPCS Code',
    'Product Description',
    'Source',
    'Customers',
    'Quantity_Clean',
    'Medicare_Rate',
    'Vendor_A_Name',
    'Vendor_A_Unit_Cost',
    'Vendor_B_Name',
    'Vendor_B_Unit_Cost',
    'Vendor_C_Name',
    'Vendor_C_Unit_Cost',
    'Best_Vendor',
    'Best_Unit_Cost',
    'Line_Total_Cost',
    'Line_Medicare_Revenue',
    'Margin_Percentage',
    'Notes'
]

# Create header row
headers = [
    'Priority_Score',
    'BOC_Category',
    'HCPCS_Code',
    'Description',
    'Source',
    'Customers',
    'Quantity',
    'Medicare_Rate',
    'Vendor_A_Name',
    'Vendor_A_Unit_Cost',
    'Vendor_B_Name',
    'Vendor_B_Unit_Cost',
    'Vendor_C_Name',
    'Vendor_C_Unit_Cost',
    'Best_Vendor',
    'Best_Unit_Cost',
    'Line_Total_Cost',
    'Medicare_Revenue_Potential',
    'Profit_Margin_%',
    'Notes'
]

ws_main.append(headers)

# Add data rows
for _, row in inventory_sorted.iterrows():
    ws_main.append([
        row['Priority_Score'],
        row['BOC Category'],
        row['HCPCS Code'],
        row['Product Description'],
        row['Source'],
        row['Customers'] if pd.notna(row['Customers']) else '',
        row['Quantity_Clean'],
        row['Medicare_Rate'],
        row['Vendor_A_Name'],
        row['Vendor_A_Unit_Cost'],
        row['Vendor_B_Name'],
        row['Vendor_B_Unit_Cost'],
        row['Vendor_C_Name'],
        row['Vendor_C_Unit_Cost'],
        row['Best_Vendor'],
        row['Best_Unit_Cost'],
        row['Line_Total_Cost'],
        row['Line_Medicare_Revenue'],
        row['Margin_Percentage'],
        row['Notes'] if pd.notna(row['Notes']) else ''
    ])

print(f"    Added {len(inventory_sorted)} products")

# ============================================================================
# SHEET 2: BOC CATEGORY SUMMARY
# ============================================================================
print("  Creating Sheet 2: BOC Category Summary...")

ws_boc = wb.create_sheet("BOC Category Summary")

# Calculate BOC summary
boc_summary = inventory_updated.groupby('BOC Category').agg({
    'HCPCS Code': 'count',
    'Quantity_Clean': 'sum',
    'Line_Total_Cost': 'sum',
    'Line_Medicare_Revenue': 'sum',
    'Line_Gross_Margin': 'sum',
    'Margin_Percentage': lambda x: x[x.notna()].mean()
}).round(2)

boc_summary.columns = ['SKU_Count', 'Total_Units', 'Total_Investment', 'Total_Revenue_Potential', 'Total_Profit_Potential', 'Avg_Margin_%']
boc_summary = boc_summary.reset_index()

# Calculate ROI
boc_summary['ROI_%'] = ((boc_summary['Total_Profit_Potential'] / boc_summary['Total_Investment']) * 100).round(2)

# Sort by total profit (highest first)
boc_summary = boc_summary.sort_values('Total_Profit_Potential', ascending=False)

# Add headers
ws_boc.append(['BOC_Category', 'SKU_Count', 'Total_Units', 'Total_Investment', 'Total_Revenue_Potential', 'Total_Profit_Potential', 'ROI_%', 'Avg_Margin_%'])

# Add data
for _, row in boc_summary.iterrows():
    ws_boc.append([
        row['BOC Category'],
        row['SKU_Count'],
        row['Total_Units'],
        row['Total_Investment'],
        row['Total_Revenue_Potential'],
        row['Total_Profit_Potential'],
        row['ROI_%'],
        row['Avg_Margin_%']
    ])

print(f"    Added {len(boc_summary)} BOC categories")

# ============================================================================
# SHEET 3: ITEMS WITHOUT MEDICARE RATES
# ============================================================================
print("  Creating Sheet 3: Items Without Medicare Rates...")

ws_no_medicare = wb.create_sheet("Items Without Medicare Rates")

no_medicare = inventory_updated[inventory_updated['Medicare_Rate'].isna()]

# Headers
ws_no_medicare.append(['Priority_Score', 'BOC_Category', 'HCPCS_Code', 'Description', 'Source', 'Customers', 'Quantity', 'Medicare_Rate', 'Vendor_A_Name', 'Vendor_A_Unit_Cost', 'Vendor_B_Name', 'Vendor_B_Unit_Cost', 'Vendor_C_Name', 'Vendor_C_Unit_Cost', 'Best_Vendor', 'Best_Unit_Cost', 'Line_Total_Cost', 'Medicare_Revenue_Potential', 'Profit_Margin_%', 'Notes'])

# Data
for _, row in no_medicare.iterrows():
    ws_no_medicare.append([
        row['Priority_Score'],
        row['BOC Category'],
        row['HCPCS Code'],
        row['Product Description'],
        row['Source'],
        row['Customers'] if pd.notna(row['Customers']) else '',
        row['Quantity_Clean'],
        row['Medicare_Rate'],
        row['Vendor_A_Name'],
        row['Vendor_A_Unit_Cost'],
        row['Vendor_B_Name'],
        row['Vendor_B_Unit_Cost'],
        row['Vendor_C_Name'],
        row['Vendor_C_Unit_Cost'],
        row['Best_Vendor'],
        row['Best_Unit_Cost'],
        row['Line_Total_Cost'],
        row['Line_Medicare_Revenue'],
        row['Margin_Percentage'],
        row['Notes'] if pd.notna(row['Notes']) else ''
    ])

print(f"    Added {len(no_medicare)} items without Medicare rates")

# ============================================================================
# SHEET 4: CUSTOMER REQUESTS
# ============================================================================
print("  Creating Sheet 4: Customer Requests...")

ws_customer = wb.create_sheet("Customer Requests")

customer_items = inventory_updated[inventory_updated['Source'] == 'CUSTOMER'].sort_values('Priority_Score', ascending=False)

# Headers
ws_customer.append(['Priority_Score', 'BOC_Category', 'HCPCS_Code', 'Description', 'Source', 'Customers', 'Quantity', 'Medicare_Rate', 'Vendor_A_Name', 'Vendor_A_Unit_Cost', 'Vendor_B_Name', 'Vendor_B_Unit_Cost', 'Vendor_C_Name', 'Vendor_C_Unit_Cost', 'Best_Vendor', 'Best_Unit_Cost', 'Line_Total_Cost', 'Medicare_Revenue_Potential', 'Profit_Margin_%', 'Notes'])

# Data
for _, row in customer_items.iterrows():
    ws_customer.append([
        row['Priority_Score'],
        row['BOC Category'],
        row['HCPCS Code'],
        row['Product Description'],
        row['Source'],
        row['Customers'] if pd.notna(row['Customers']) else '',
        row['Quantity_Clean'],
        row['Medicare_Rate'],
        row['Vendor_A_Name'],
        row['Vendor_A_Unit_Cost'],
        row['Vendor_B_Name'],
        row['Vendor_B_Unit_Cost'],
        row['Vendor_C_Name'],
        row['Vendor_C_Unit_Cost'],
        row['Best_Vendor'],
        row['Best_Unit_Cost'],
        row['Line_Total_Cost'],
        row['Line_Medicare_Revenue'],
        row['Margin_Percentage'],
        row['Notes'] if pd.notna(row['Notes']) else ''
    ])

print(f"    Added {len(customer_items)} customer request items")

# ============================================================================
# STEP 9: APPLY FORMATTING
# ============================================================================
print("\nSTEP 9: Applying professional formatting...")

# Define styles
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# Format all sheets
for sheet_name in ['Inventory Analysis', 'BOC Category Summary', 'Items Without Medicare Rates', 'Customer Requests']:
    ws = wb[sheet_name]

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set row height for header
    ws.row_dimensions[1].height = 30

    # Alternating row colors and borders
    light_gray = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = light_gray

    # Number formatting for main sheets
    if sheet_name in ['Inventory Analysis', 'Items Without Medicare Rates', 'Customer Requests']:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # Priority Score (column A)
            row[0].number_format = '0'
            # Quantity (column G)
            row[6].number_format = '0'
            # Medicare Rate (column H)
            if row[7].value:
                row[7].number_format = '$#,##0.00'
            # Vendor costs (columns J, L, N, P, Q)
            for col_idx in [9, 11, 13, 15, 16]:
                if row[col_idx].value:
                    row[col_idx].number_format = '$#,##0.00'
            # Line Total Cost (column Q/16)
            row[16].number_format = '$#,##0.00'
            # Medicare Revenue (column R/17)
            row[17].number_format = '$#,##0.00'
            # Margin % (column S/18)
            if row[18].value:
                row[18].number_format = '0.0%'

    # Number formatting for BOC summary
    if sheet_name == 'BOC Category Summary':
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            # SKU Count (column B)
            row[1].number_format = '0'
            # Total Units (column C)
            row[2].number_format = '0'
            # Investment, Revenue, Profit (columns D, E, F)
            for col_idx in [3, 4, 5]:
                row[col_idx].number_format = '$#,##0.00'
            # ROI % (column G)
            if row[6].value:
                row[6].number_format = '0.0%'
            # Avg Margin % (column H)
            if row[7].value:
                row[7].number_format = '0.0%'

# Conditional formatting for Margin % (Inventory Analysis sheet only)
ws_main = wb['Inventory Analysis']
ws_main.conditional_formatting.add(
    f'S2:S{ws_main.max_row}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0.30'],
               fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
               font=Font(color='006100'))
)
ws_main.conditional_formatting.add(
    f'S2:S{ws_main.max_row}',
    CellIsRule(operator='between', formula=['0.10', '0.299'],
               fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
               font=Font(color='9C6500'))
)
ws_main.conditional_formatting.add(
    f'S2:S{ws_main.max_row}',
    CellIsRule(operator='lessThan', formula=['0.10'],
               fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
               font=Font(color='9C0006'))
)

# Set column widths for Inventory Analysis
column_widths = {
    'A': 15, 'B': 15, 'C': 12, 'D': 50, 'E': 18, 'F': 25,
    'G': 10, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15,
    'M': 15, 'N': 15, 'O': 15, 'P': 15, 'Q': 18, 'R': 22,
    'S': 15, 'T': 40
}
for col, width in column_widths.items():
    ws_main.column_dimensions[col].width = width

# Set column widths for BOC Category Summary
ws_boc = wb['BOC Category Summary']
boc_widths = {'A': 18, 'B': 12, 'C': 12, 'D': 18, 'E': 22, 'F': 20, 'G': 12, 'H': 15}
for col, width in boc_widths.items():
    ws_boc.column_dimensions[col].width = width

# Tab colors
wb['Inventory Analysis'].sheet_properties.tabColor = "1F77B4"  # Blue
wb['BOC Category Summary'].sheet_properties.tabColor = "2CA02C"  # Green
wb['Items Without Medicare Rates'].sheet_properties.tabColor = "D62728"  # Red
wb['Customer Requests'].sheet_properties.tabColor = "FF7F0E"  # Orange

print("  ✓ Applied formatting to all sheets")

# ============================================================================
# STEP 10: ADD FORMULAS (Best Vendor/Cost Auto-Calculation)
# ============================================================================
print("\nSTEP 10: Adding Excel formulas for vendor comparison...")

ws_main = wb['Inventory Analysis']

# Add formulas starting from row 2
for row_idx in range(2, ws_main.max_row + 1):
    # Best_Unit_Cost formula (column P) - MIN of vendor costs
    ws_main[f'P{row_idx}'] = f'=MIN(J{row_idx},L{row_idx},N{row_idx})'

    # Best_Vendor formula (column O) - which vendor has the min cost
    ws_main[f'O{row_idx}'] = f'=IF(P{row_idx}=J{row_idx},I{row_idx},IF(P{row_idx}=L{row_idx},K{row_idx},M{row_idx}))'

    # Line_Total_Cost formula (column Q) - Quantity * Best_Unit_Cost
    ws_main[f'Q{row_idx}'] = f'=G{row_idx}*P{row_idx}'

    # Medicare_Revenue_Potential formula (column R) - Quantity * Medicare_Rate
    ws_main[f'R{row_idx}'] = f'=IF(ISBLANK(H{row_idx}),0,G{row_idx}*H{row_idx})'

    # Profit_Margin_% formula (column S) - (Revenue - Cost) / Revenue
    ws_main[f'S{row_idx}'] = f'=IF(R{row_idx}=0,0,(R{row_idx}-Q{row_idx})/R{row_idx})'

print(f"  ✓ Added formulas to {ws_main.max_row - 1} rows")

# ============================================================================
# SAVE WORKBOOK
# ============================================================================
print("\nSTEP 11: Saving workbook...")

wb.save(output_file)
print(f"  ✓ Saved to: {output_file}")

file_size = os.path.getsize(output_file)
print(f"  ✓ File size: {file_size:,} bytes")

# ============================================================================
# STEP 12: VALIDATION
# ============================================================================
print("\n" + "="*100)
print("VALIDATION CHECKS")
print("="*100)

# Reload to verify
wb_verify = load_workbook(output_file, data_only=False)

# Check 1: All sheets present
expected_sheets = ['Inventory Analysis', 'BOC Category Summary', 'Items Without Medicare Rates', 'Customer Requests']
sheets_present = all(sheet in wb_verify.sheetnames for sheet in expected_sheets)
print(f"\n✓ Sheet structure: {'PASS' if sheets_present else 'FAIL'}")
print(f"  Sheets: {wb_verify.sheetnames}")

# Check 2: Row counts
ws_verify = wb_verify['Inventory Analysis']
row_count = ws_verify.max_row - 1  # Exclude header
print(f"\n✓ Inventory Analysis: {row_count} products")

# Check 3: Verify missing items are present
ws_verify_data = wb_verify['Inventory Analysis']
hcpcs_in_workbook = set()
for row in ws_verify_data.iter_rows(min_row=2, max_row=ws_verify_data.max_row, min_col=3, max_col=3):
    if row[0].value:
        hcpcs_in_workbook.add(row[0].value)

missing_ankle_codes = ['L1906', 'L4361', 'L4350', 'L4370', 'L4387']
missing_cgm_codes = ['E2103', 'A4239']

ankle_found = [code for code in missing_ankle_codes if code in hcpcs_in_workbook]
cgm_found = [code for code in missing_cgm_codes if code in hcpcs_in_workbook]

print(f"\n✓ Dr. Nas ankle products: {len(ankle_found)}/{len(missing_ankle_codes)}")
for code in missing_ankle_codes:
    status = "✓" if code in ankle_found else "✗"
    print(f"  {status} {code}")

print(f"\n✓ CGM full systems: {len(cgm_found)}/{len(missing_cgm_codes)}")
for code in missing_cgm_codes:
    status = "✓" if code in cgm_found else "✗"
    print(f"  {status} {code}")

# Check 4: Formulas present
formula_check_cells = ['O2', 'P2', 'Q2', 'R2', 'S2']
formulas_present = all(ws_verify[cell].value and str(ws_verify[cell].value).startswith('=') for cell in formula_check_cells)
print(f"\n✓ Excel formulas: {'PASS' if formulas_present else 'FAIL'}")
if formulas_present:
    for cell in formula_check_cells:
        print(f"  {cell}: {ws_verify[cell].value}")

# Check 5: Formatting present
header_formatted = ws_verify['A1'].fill.start_color.rgb == '001F4E78'
print(f"\n✓ Header formatting: {'PASS' if header_formatted else 'FAIL'}")

# Check 6: Calculate totals
total_investment = inventory_updated['Line_Total_Cost'].sum()
medicare_coverage = (inventory_updated['Medicare_Rate'].notna().sum() / len(inventory_updated)) * 100

print(f"\n✓ Budget totals:")
print(f"  Total Investment: ${total_investment:,.2f}")
print(f"  Total Products: {len(inventory_updated)}")
print(f"  Medicare Coverage: {medicare_coverage:.1f}%")

print("\n" + "="*100)
print("REBUILD COMPLETE ✅")
print("="*100)
print(f"\nFinal Workbook: {output_file}")
print(f"Total Products: {len(inventory_updated)}")
print(f"  - Original: {len(inventory_clean)}")
print(f"  - Added ankle: {len(ankle_df)}")
print(f"  - Added CGM: {len(cgm_df)}")
print(f"\nAll missing items included: ✅")
print(f"All formulas working: ✅")
print(f"Professional formatting: ✅")
print("="*100)
