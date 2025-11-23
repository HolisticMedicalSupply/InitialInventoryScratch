#!/usr/bin/env python3
"""
Rebuild Final VGM Vendor Analysis Workbook with ALL TIERS
- Uses MASTER_INVENTORY_PLAN_COMPLETE_ALL_TIERS.csv (383 products)
- 27 BOC categories (79.4% coverage)
- All formulas, formatting, and 4 sheets
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

print("="*100)
print("REBUILDING FINAL VGM VENDOR ANALYSIS WORKBOOK - COMPLETE ALL TIERS")
print("="*100)

# ============================================================================
# STEP 1: LOAD COMPREHENSIVE INVENTORY
# ============================================================================
print("\nSTEP 1: Loading comprehensive inventory plan...")

df = pd.read_csv("/home/user/InitialInventoryScratch/MASTER_INVENTORY_PLAN_COMPLETE_ALL_TIERS.csv")
print(f"  Total products: {len(df)}")
print(f"  Total BOC categories: {df['BOC_Category'].nunique()}")
print(f"  Total investment: ${df['Total_Cost'].sum():,.2f}")

# ============================================================================
# STEP 2: SORT BY PRIORITY SCORE (HIGHEST FIRST)
# ============================================================================
print("\nSTEP 2: Sorting by priority score...")

df = df.sort_values('Priority_Score', ascending=False).reset_index(drop=True)
print(f"  Highest priority: {df.iloc[0]['Priority_Score']} - {df.iloc[0]['HCPCS_Code']} ({df.iloc[0]['Description']})")
print(f"  Lowest priority: {df.iloc[-1]['Priority_Score']} - {df.iloc[-1]['HCPCS_Code']} ({df.iloc[-1]['Description']})")

# ============================================================================
# STEP 3: CREATE WORKBOOK WITH 4 SHEETS
# ============================================================================
print("\nSTEP 3: Creating workbook with 4 sheets...")

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Sheet 1: Main Inventory Analysis
ws_main = wb.create_sheet("Inventory Analysis", 0)

# Sheet 2: BOC Category Summary
ws_boc = wb.create_sheet("BOC Category Summary", 1)

# Sheet 3: Items Without Medicare Rates
ws_no_medicare = wb.create_sheet("Items Without Medicare Rates", 2)

# Sheet 4: Customer Requests
ws_customers = wb.create_sheet("Customer Requests", 3)

print(f"  ✓ Created 4 sheets")

# ============================================================================
# SHEET 1: MAIN INVENTORY ANALYSIS
# ============================================================================
print("\nSTEP 4: Building Sheet 1 - Inventory Analysis...")

# Define headers
headers = [
    'HCPCS Code', 'BOC Category', 'Description', 'Quantity',
    'Medicare Allowable Rate',
    'Vendor A Name', 'Vendor A Unit Cost',
    'Vendor B Name', 'Vendor B Unit Cost',
    'Vendor C Name', 'Vendor C Unit Cost',
    'Best Vendor', 'Best Unit Cost', 'Line Total Cost',
    'Medicare Revenue', 'Profit Margin %',
    'Priority', 'Source', 'Customer'
]

# Write headers
ws_main.append(headers)

# Style header row
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
for cell in ws_main[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Freeze header row
ws_main.freeze_panes = "A2"

# Write data rows with formulas
row_idx = 2
for idx, row in df.iterrows():
    # Extract data
    hcpcs = row['HCPCS_Code']
    boc = row['BOC_Category']
    desc = row['Description']
    qty = int(row['Quantity'])
    medicare_rate = row['Medicare_Rate'] if pd.notna(row['Medicare_Rate']) else None
    priority = int(row['Priority_Score'])
    source = row['Source']
    customer = row['Customers'] if pd.notna(row['Customers']) else ""

    # Write data
    ws_main.append([
        hcpcs,
        boc,
        desc,
        qty,
        medicare_rate,
        '',  # Vendor A Name
        '',  # Vendor A Unit Cost (J)
        '',  # Vendor B Name
        '',  # Vendor B Unit Cost (L)
        '',  # Vendor C Name
        '',  # Vendor C Unit Cost (N)
        '',  # Best Vendor (O) - FORMULA
        '',  # Best Unit Cost (P) - FORMULA
        '',  # Line Total Cost (Q) - FORMULA
        '',  # Medicare Revenue (R) - FORMULA
        '',  # Profit Margin % (S) - FORMULA
        priority,
        source,
        customer
    ])

    # Add formulas in columns O, P, Q, R, S
    # O: Best_Vendor = IF(P=J, Vendor_A, IF(P=L, Vendor_B, Vendor_C))
    ws_main[f'O{row_idx}'] = f'=IF(P{row_idx}=J{row_idx},I{row_idx},IF(P{row_idx}=L{row_idx},K{row_idx},M{row_idx}))'

    # P: Best_Unit_Cost = MIN(J, L, N)
    ws_main[f'P{row_idx}'] = f'=MIN(J{row_idx},L{row_idx},N{row_idx})'

    # Q: Line_Total_Cost = Quantity * Best_Unit_Cost
    ws_main[f'Q{row_idx}'] = f'=G{row_idx}*P{row_idx}'

    # R: Medicare_Revenue = IF(ISBLANK(Medicare_Rate), 0, Quantity * Medicare_Rate)
    ws_main[f'R{row_idx}'] = f'=IF(ISBLANK(H{row_idx}),0,G{row_idx}*H{row_idx})'

    # S: Profit_Margin_% = IF(Revenue=0, 0, (Revenue - Cost) / Revenue)
    ws_main[f'S{row_idx}'] = f'=IF(R{row_idx}=0,0,(R{row_idx}-Q{row_idx})/R{row_idx})'

    row_idx += 1

print(f"  ✓ Added {len(df)} products with formulas")

# Apply conditional formatting to Profit Margin % column (S)
# Green: >30%, Yellow: 10-30%, Red: <10%
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

ws_main.conditional_formatting.add(f'S2:S{row_idx-1}',
    CellIsRule(operator='greaterThan', formula=['0.3'], stopIfTrue=True, fill=green_fill))
ws_main.conditional_formatting.add(f'S2:S{row_idx-1}',
    CellIsRule(operator='between', formula=['0.1', '0.3'], stopIfTrue=True, fill=yellow_fill))
ws_main.conditional_formatting.add(f'S2:S{row_idx-1}',
    CellIsRule(operator='lessThan', formula=['0.1'], stopIfTrue=True, fill=red_fill))

print(f"  ✓ Applied conditional formatting")

# Set column widths
column_widths = {
    'A': 12,  # HCPCS
    'B': 12,  # BOC
    'C': 50,  # Description
    'D': 10,  # Quantity
    'E': 12,  # Medicare Rate
    'F': 15,  # Vendor A Name
    'G': 12,  # Vendor A Cost
    'H': 15,  # Vendor B Name
    'I': 12,  # Vendor B Cost
    'J': 15,  # Vendor C Name
    'K': 12,  # Vendor C Cost
    'L': 15,  # Best Vendor
    'M': 12,  # Best Cost
    'N': 12,  # Line Total
    'O': 15,  # Medicare Revenue
    'P': 12,  # Profit Margin %
    'Q': 10,  # Priority
    'R': 18,  # Source
    'S': 20,  # Customer
}

for col, width in column_widths.items():
    ws_main.column_dimensions[col].width = width

# Format currency and percentage columns
for row in range(2, row_idx):
    ws_main[f'E{row}'].number_format = '$#,##0.00'  # Medicare Rate
    ws_main[f'G{row}'].number_format = '$#,##0.00'  # Vendor A Cost
    ws_main[f'I{row}'].number_format = '$#,##0.00'  # Vendor B Cost
    ws_main[f'K{row}'].number_format = '$#,##0.00'  # Vendor C Cost
    ws_main[f'M{row}'].number_format = '$#,##0.00'  # Best Cost
    ws_main[f'N{row}'].number_format = '$#,##0.00'  # Line Total
    ws_main[f'O{row}'].number_format = '$#,##0.00'  # Medicare Revenue
    ws_main[f'P{row}'].number_format = '0.0%'       # Profit Margin %

print(f"  ✓ Applied number formatting")

# ============================================================================
# SHEET 2: BOC CATEGORY SUMMARY
# ============================================================================
print("\nSTEP 5: Building Sheet 2 - BOC Category Summary...")

boc_summary = df.groupby('BOC_Category').agg({
    'HCPCS_Code': 'count',
    'Total_Cost': 'sum',
    'Total_Revenue_Potential': 'sum',
    'Total_Profit_Potential': 'sum'
}).rename(columns={
    'HCPCS_Code': 'SKU_Count',
    'Total_Cost': 'Investment',
    'Total_Revenue_Potential': 'Revenue_Potential',
    'Total_Profit_Potential': 'Profit_Potential'
})

boc_summary['ROI_%'] = (boc_summary['Profit_Potential'] / boc_summary['Investment']) * 100
boc_summary['Avg_Margin_%'] = (boc_summary['Profit_Potential'] / boc_summary['Revenue_Potential']) * 100

# Sort by profit potential (highest first)
boc_summary = boc_summary.sort_values('Profit_Potential', ascending=False)

# Write to sheet
boc_headers = ['BOC Category', 'SKU Count', 'Investment', 'Revenue Potential', 'Profit Potential', 'ROI %', 'Avg Margin %']
ws_boc.append(boc_headers)

# Style header
for cell in ws_boc[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws_boc.freeze_panes = "A2"

# Write data
for boc, row in boc_summary.iterrows():
    ws_boc.append([
        boc,
        int(row['SKU_Count']),
        row['Investment'],
        row['Revenue_Potential'],
        row['Profit_Potential'],
        row['ROI_%'],
        row['Avg_Margin_%']
    ])

# Format columns
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_boc.column_dimensions[col].width = 18

for row_num in range(2, len(boc_summary) + 2):
    ws_boc[f'C{row_num}'].number_format = '$#,##0.00'
    ws_boc[f'D{row_num}'].number_format = '$#,##0.00'
    ws_boc[f'E{row_num}'].number_format = '$#,##0.00'
    ws_boc[f'F{row_num}'].number_format = '0.0%'
    ws_boc[f'G{row_num}'].number_format = '0.0%'

print(f"  ✓ Added {len(boc_summary)} BOC categories")

# ============================================================================
# SHEET 3: ITEMS WITHOUT MEDICARE RATES
# ============================================================================
print("\nSTEP 6: Building Sheet 3 - Items Without Medicare Rates...")

no_medicare = df[df['Medicare_Rate'].isna()].copy()

# Write headers
no_medicare_headers = ['HCPCS Code', 'BOC Category', 'Description', 'Quantity', 'Est Cost', 'Total Cost', 'Notes']
ws_no_medicare.append(no_medicare_headers)

# Style header
for cell in ws_no_medicare[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws_no_medicare.freeze_panes = "A2"

# Write data
for idx, row in no_medicare.iterrows():
    notes = "Private pay item - no Medicare coverage"
    if row['BOC_Category'] == 'DM06':
        notes = "CGM receivers - typically private pay or through sensor benefit"
    elif row['BOC_Category'] in ['PD09', 'S04']:
        notes = "High-volume consumable - may have alternative coverage"

    ws_no_medicare.append([
        row['HCPCS_Code'],
        row['BOC_Category'],
        row['Description'],
        int(row['Quantity']),
        row['Estimated_Unit_Cost'],
        row['Total_Cost'],
        notes
    ])

# Format columns
ws_no_medicare.column_dimensions['A'].width = 12
ws_no_medicare.column_dimensions['B'].width = 12
ws_no_medicare.column_dimensions['C'].width = 50
ws_no_medicare.column_dimensions['D'].width = 10
ws_no_medicare.column_dimensions['E'].width = 12
ws_no_medicare.column_dimensions['F'].width = 12
ws_no_medicare.column_dimensions['G'].width = 50

for row_num in range(2, len(no_medicare) + 2):
    ws_no_medicare[f'E{row_num}'].number_format = '$#,##0.00'
    ws_no_medicare[f'F{row_num}'].number_format = '$#,##0.00'

print(f"  ✓ Added {len(no_medicare)} items without Medicare rates")

# ============================================================================
# SHEET 4: CUSTOMER REQUESTS
# ============================================================================
print("\nSTEP 7: Building Sheet 4 - Customer Requests...")

customer_items = df[df['Source'] == 'CUSTOMER'].copy()
customer_items = customer_items.sort_values('Priority_Score', ascending=False)

# Write headers
customer_headers = ['HCPCS Code', 'BOC Category', 'Description', 'Quantity', 'Est Cost', 'Total Cost', 'Customer', 'Priority']
ws_customers.append(customer_headers)

# Style header
for cell in ws_customers[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws_customers.freeze_panes = "A2"

# Write data
for idx, row in customer_items.iterrows():
    ws_customers.append([
        row['HCPCS_Code'],
        row['BOC_Category'],
        row['Description'],
        int(row['Quantity']),
        row['Estimated_Unit_Cost'],
        row['Total_Cost'],
        row['Customers'] if pd.notna(row['Customers']) else '',
        int(row['Priority_Score'])
    ])

# Format columns
ws_customers.column_dimensions['A'].width = 12
ws_customers.column_dimensions['B'].width = 12
ws_customers.column_dimensions['C'].width = 50
ws_customers.column_dimensions['D'].width = 10
ws_customers.column_dimensions['E'].width = 12
ws_customers.column_dimensions['F'].width = 12
ws_customers.column_dimensions['G'].width = 25
ws_customers.column_dimensions['H'].width = 10

for row_num in range(2, len(customer_items) + 2):
    ws_customers[f'E{row_num}'].number_format = '$#,##0.00'
    ws_customers[f'F{row_num}'].number_format = '$#,##0.00'

print(f"  ✓ Added {len(customer_items)} customer-requested items")

# ============================================================================
# SAVE WORKBOOK
# ============================================================================
output_file = "/home/user/InitialInventoryScratch/Holistic_Medical_VGM_Vendor_Analysis_COMPLETE_ALL_TIERS_2025-11-22.xlsx"
wb.save(output_file)

print(f"\n{'='*100}")
print(f"✅ WORKBOOK COMPLETE")
print(f"{'='*100}")
print(f"  File: {output_file}")
print(f"  Total Products: {len(df)}")
print(f"  BOC Categories: {df['BOC_Category'].nunique()}")
print(f"  Total Investment: ${df['Total_Cost'].sum():,.2f}")
print(f"  Revenue Potential: ${df['Total_Revenue_Potential'].sum():,.2f}")
print(f"  Profit Potential: ${df['Total_Profit_Potential'].sum():,.2f}")
print(f"  Customer Items: {len(customer_items)}")
print(f"  Non-Medicare Items: {len(no_medicare)}")
print(f"{'='*100}")
