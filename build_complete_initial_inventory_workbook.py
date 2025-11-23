#!/usr/bin/env python3
"""
Build COMPLETE Initial Inventory Plan Workbook
Includes ALL items from:
- Customer Needs PDF
- Launch Inventory PDF (all 9 tiers)
- Existing MASTER_INVENTORY_PLAN_COMPLETE_ALL_TIERS.csv
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

print("="*80)
print("BUILDING COMPLETE INITIAL INVENTORY PLAN WORKBOOK")
print("="*80)

# Step 1: Load Medicare rates from normalized structure
print("\n[1/8] Loading Medicare rates...")
medicare_rates_df = pd.read_excel('Medicare_Rates_Normalized_Structure_Validated.xlsx', sheet_name='Medicare Rates - Normalized')
# Get average rate per HCPCS code (some codes have multiple rates for different modifiers/tiers)
medicare_rates_grouped = medicare_rates_df.groupby('HCPCS Code')['Rate ($)'].mean().to_dict()
medicare_rates_dict = medicare_rates_grouped
print(f"   Loaded {len(medicare_rates_dict)} Medicare rates")

# Step 2: Load existing complete inventory CSV
print("\n[2/8] Loading existing inventory...")
existing_df = pd.read_csv('MASTER_INVENTORY_PLAN_COMPLETE_ALL_TIERS.csv')
print(f"   Loaded {len(existing_df)} existing items")

# Step 3: Define ALL missing items from analysis
print("\n[3/8] Adding missing critical items...")

missing_items = []

# === HIGH PRIORITY CUSTOMER REQUESTS ===
# NOTE: CPAP equipment, baby formula (oral), orthopedic shoes, vacuum equipment
# are either not DME-billable or outside current BOC scope - documented as exclusions

# Shower Chair (Customer delivery report)
missing_items.extend([
    {'HCPCS': 'E0240', 'BOC': 'DM02', 'Desc': 'Bath/shower chair, with or without wheels, any size', 'Qty': 5, 'Unit_Cost': 65.0, 'Priority': 80, 'Source': 'CUSTOMER', 'Customer': 'WALTERS'},
])

# Wheelchair Cushions (CRITICAL - needed for all wheelchair sales)
missing_items.extend([
    {'HCPCS': 'E2601', 'BOC': 'M06', 'Desc': 'General use wheelchair seat cushion, width less than 22 inches, any depth', 'Qty': 8, 'Unit_Cost': 45.0, 'Priority': 90, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'E2602', 'BOC': 'M06', 'Desc': 'General use wheelchair seat cushion, width 22 inches or greater, any depth', 'Qty': 8, 'Unit_Cost': 55.0, 'Priority': 90, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'E2603', 'BOC': 'M06', 'Desc': 'Skin protection wheelchair seat cushion, width less than 22 inches, any depth', 'Qty': 8, 'Unit_Cost': 85.0, 'Priority': 85, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Urological Supplies - Catheters (TIER 3 - HIGH VOLUME)
missing_items.extend([
    {'HCPCS': 'A4338', 'BOC': 'PD09', 'Desc': 'Indwelling catheter, Foley type, two-way latex with coating, each', 'Qty': 30, 'Unit_Cost': 3.50, 'Priority': 70, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4340', 'BOC': 'PD09', 'Desc': 'Indwelling catheter, Foley type, two-way, all silicone, each', 'Qty': 20, 'Unit_Cost': 5.50, 'Priority': 70, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4344', 'BOC': 'PD09', 'Desc': 'Indwelling catheter, Foley type, two-way, all silicone, pediatric, each', 'Qty': 10, 'Unit_Cost': 6.00, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4346', 'BOC': 'PD09', 'Desc': 'Indwelling catheter, Foley type, three-way for continuous irrigation, each', 'Qty': 10, 'Unit_Cost': 7.50, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4349', 'BOC': 'PD09', 'Desc': 'Male external catheter, with or without adhesive, disposable, each', 'Qty': 40, 'Unit_Cost': 1.25, 'Priority': 65, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4351', 'BOC': 'PD09', 'Desc': 'Intermittent urinary catheter, straight tip, with or without coating, each', 'Qty': 30, 'Unit_Cost': 2.50, 'Priority': 65, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4353', 'BOC': 'PD09', 'Desc': 'Intermittent urinary catheter, with insertion supplies', 'Qty': 20, 'Unit_Cost': 3.75, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Urological Drainage Bags
missing_items.extend([
    {'HCPCS': 'A4357', 'BOC': 'PD09', 'Desc': 'Bedside drainage bag, day or night, with or without anti-reflux device, with or without tube, each', 'Qty': 30, 'Unit_Cost': 4.50, 'Priority': 70, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4358', 'BOC': 'PD09', 'Desc': 'Urinary leg bag, vinyl, with or without tube, each', 'Qty': 40, 'Unit_Cost': 3.25, 'Priority': 70, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Catheter Irrigation & Insertion Supplies
missing_items.extend([
    {'HCPCS': 'A4320', 'BOC': 'PD09', 'Desc': 'Irrigation tray, bladder or urethral, complete set', 'Qty': 20, 'Unit_Cost': 8.50, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4322', 'BOC': 'PD09', 'Desc': 'Irrigation syringe, bulb or piston, each', 'Qty': 15, 'Unit_Cost': 2.25, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4310', 'BOC': 'PD09', 'Desc': 'Insertion tray without drainage bag and without catheter', 'Qty': 10, 'Unit_Cost': 5.50, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4316', 'BOC': 'PD09', 'Desc': 'Insertion tray with drainage bag, with indwelling catheter', 'Qty': 10, 'Unit_Cost': 12.50, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Non-disposable underpads
missing_items.extend([
    {'HCPCS': 'A4553', 'BOC': 'PD09', 'Desc': 'Non-disposable underpads, all sizes', 'Qty': 20, 'Unit_Cost': 15.00, 'Priority': 65, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Diabetes Accessories
missing_items.extend([
    {'HCPCS': 'A4256', 'BOC': 'DM06', 'Desc': 'Normal, low and high calibrator solution/chips', 'Qty': 30, 'Unit_Cost': 8.00, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4258', 'BOC': 'DM06', 'Desc': 'Spring-powered device for lancet, each', 'Qty': 25, 'Unit_Cost': 5.00, 'Priority': 65, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# TENS Supplies
missing_items.extend([
    {'HCPCS': 'A4557', 'BOC': 'DM22', 'Desc': 'Lead wires, (e.g., apnea monitor), pair', 'Qty': 15, 'Unit_Cost': 12.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4556', 'BOC': 'DM22', 'Desc': 'Electrodes, (e.g., apnea monitor), per pair', 'Qty': 20, 'Unit_Cost': 8.50, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4630', 'BOC': 'DM22', 'Desc': 'Replacement batteries, medically necessary, transcutaneous electrical nerve stimulator, owned by patient', 'Qty': 30, 'Unit_Cost': 6.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Nebulizer Supplies (expanded)
missing_items.extend([
    {'HCPCS': 'A7004', 'BOC': 'R07', 'Desc': 'Small volume nebulizer, disposable, used with aerosol compressor', 'Qty': 25, 'Unit_Cost': 3.50, 'Priority': 60, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7005', 'BOC': 'R07', 'Desc': 'Administration set, with small volume nonfiltered pneumatic nebulizer, disposable', 'Qty': 20, 'Unit_Cost': 4.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7006', 'BOC': 'R07', 'Desc': 'Administration set, with small volume filtered pneumatic nebulizer', 'Qty': 15, 'Unit_Cost': 5.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7010', 'BOC': 'R07', 'Desc': 'Corrugated tubing, disposable, used with large volume nebulizer, 100 feet', 'Qty': 10, 'Unit_Cost': 12.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7014', 'BOC': 'R07', 'Desc': 'Filter, disposable, used with aerosol compressor or ultrasonic generator', 'Qty': 25, 'Unit_Cost': 2.50, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Suction Supplies
missing_items.extend([
    {'HCPCS': 'A7002', 'BOC': 'DM29', 'Desc': 'Tubing, used with suction pump, each', 'Qty': 20, 'Unit_Cost': 3.50, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Pneumatic Compression Supplies
missing_items.extend([
    {'HCPCS': 'A4600', 'BOC': 'DM18', 'Desc': 'Sleeve for intermittent limb compression device, replacement only, each', 'Qty': 15, 'Unit_Cost': 45.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'E0667', 'BOC': 'DM18', 'Desc': 'Segmental pneumatic appliance for use with pneumatic compressor, full arm', 'Qty': 3, 'Unit_Cost': 125.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'E0668', 'BOC': 'DM18', 'Desc': 'Segmental pneumatic appliance for use with pneumatic compressor, full leg', 'Qty': 3, 'Unit_Cost': 135.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'E0669', 'BOC': 'DM18', 'Desc': 'Segmental pneumatic appliance for use with pneumatic compressor, half leg', 'Qty': 3, 'Unit_Cost': 110.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'E0673', 'BOC': 'DM18', 'Desc': 'Segmental gradient pressure pneumatic appliance, half leg', 'Qty': 3, 'Unit_Cost': 120.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Cold Therapy
missing_items.extend([
    {'HCPCS': 'E0236', 'BOC': 'DM08', 'Desc': 'Pump for water circulating pad', 'Qty': 3, 'Unit_Cost': 95.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Osteogenesis Stimulator
missing_items.extend([
    {'HCPCS': 'E0748', 'BOC': 'DM17', 'Desc': 'Osteogenesis stimulator, electrical, non-invasive, spinal applications', 'Qty': 1, 'Unit_Cost': 850.00, 'Priority': 45, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A4559', 'BOC': 'DM17', 'Desc': 'Coupling gel or paste, for use with ultrasound device, per oz', 'Qty': 10, 'Unit_Cost': 4.50, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Traction Equipment
missing_items.extend([
    {'HCPCS': 'E0856', 'BOC': 'DM21', 'Desc': 'Cervical traction equipment not requiring additional stand or frame', 'Qty': 1, 'Unit_Cost': 185.00, 'Priority': 45, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'E0900', 'BOC': 'DM21', 'Desc': 'Traction stand, free standing, extremity traction', 'Qty': 1, 'Unit_Cost': 275.00, 'Priority': 40, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Orthotics (missing codes)
missing_items.extend([
    {'HCPCS': 'L1833', 'BOC': 'OR03', 'Desc': 'Knee orthosis, adjustable knee joints, other than condylar type, prefabricated, off-the-shelf', 'Qty': 3, 'Unit_Cost': 120.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'L0628', 'BOC': 'OR03', 'Desc': 'Lumbar-sacral orthosis, flexible, provides lumbar support, posterior extends from L-1 to below L-5 vertebra, produces intracavitary pressure, prefabricated, off-the-shelf', 'Qty': 3, 'Unit_Cost': 85.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'L3710', 'BOC': 'OR03', 'Desc': 'Elbow orthosis, elastic with stays, prefabricated, off-the-shelf', 'Qty': 3, 'Unit_Cost': 45.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Enteral Formulas (missing)
missing_items.extend([
    {'HCPCS': 'B4152', 'BOC': 'PE03', 'Desc': 'Enteral formula, nutritionally complete, calorically dense (equal to or greater than 1.5 kcal/ml), administered through enteral feeding tube, 100 calories = 1 unit', 'Qty': 15, 'Unit_Cost': 0.60, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'B4154', 'BOC': 'PE03', 'Desc': 'Enteral formula, nutritionally complete, for special metabolic needs, excludes inherited disease of metabolism, administered through enteral feeding tube, 100 calories = 1 unit', 'Qty': 10, 'Unit_Cost': 0.75, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'B4155', 'BOC': 'PE03', 'Desc': 'Enteral formula, nutritionally incomplete/modular nutrients, administered through enteral feeding tube, 100 calories = 1 unit', 'Qty': 10, 'Unit_Cost': 0.65, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'B4162', 'BOC': 'PE03', 'Desc': 'Enteral formula, for pediatrics, nutritionally complete, calorically dense (equal to or greater than 0.7 kcal/ml), administered through enteral feeding tube, 100 calories = 1 unit', 'Qty': 5, 'Unit_Cost': 0.70, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Enteral Equipment
missing_items.extend([
    {'HCPCS': 'B4036', 'BOC': 'PE04', 'Desc': 'Enteral feeding supply kit, gravity fed, per day', 'Qty': 10, 'Unit_Cost': 12.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'B4082', 'BOC': 'PE04', 'Desc': 'Nasogastric tubing with stylet', 'Qty': 8, 'Unit_Cost': 8.50, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'B4083', 'BOC': 'PE04', 'Desc': 'Stomach tube - Levine type', 'Qty': 7, 'Unit_Cost': 7.50, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Tracheostomy Tubes (additional sizes)
missing_items.extend([
    {'HCPCS': 'A7522', 'BOC': 'PD08', 'Desc': 'Tracheostomy, inner cannula', 'Qty': 5, 'Unit_Cost': 35.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7523', 'BOC': 'PD08', 'Desc': 'Tracheostomy, shower protector, each', 'Qty': 5, 'Unit_Cost': 15.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7524', 'BOC': 'PD08', 'Desc': 'Tracheostoma stent/stud/button, each', 'Qty': 3, 'Unit_Cost': 45.00, 'Priority': 45, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7525', 'BOC': 'PD08', 'Desc': 'Tracheostomy mask, each', 'Qty': 5, 'Unit_Cost': 8.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A7526', 'BOC': 'PD08', 'Desc': 'Tracheostomy tube collar/holder, each', 'Qty': 5, 'Unit_Cost': 6.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Basic Supplies - HIGH VOLUME CONSUMABLES
missing_items.extend([
    {'HCPCS': 'A9286', 'BOC': 'S01', 'Desc': 'Hygienic item or device, disposable or non-disposable, any type, each', 'Qty': 200, 'Unit_Cost': 0.50, 'Priority': 70, 'Source': 'CUSTOMER', 'Customer': 'WALTERS'},  # Wet wipes
])

# Additional Compression Stockings (to complete product line)
missing_items.extend([
    {'HCPCS': 'A6536', 'BOC': 'S04', 'Desc': 'Gradient compression stocking, thigh length, 30-40 mmHg, custom, each', 'Qty': 3, 'Unit_Cost': 95.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A6537', 'BOC': 'S04', 'Desc': 'Gradient compression stocking, thigh length, 40 mmHg or greater, custom, each', 'Qty': 3, 'Unit_Cost': 105.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A6544', 'BOC': 'S04', 'Desc': 'Gradient compression stocking, garter belt', 'Qty': 3, 'Unit_Cost': 25.00, 'Priority': 50, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

# Additional Wound Care (specific missing codes)
missing_items.extend([
    {'HCPCS': 'A6198', 'BOC': 'S01', 'Desc': 'Alginate or other fiber gelling dressing, wound cover, sterile, pad size 16 sq. in. or less, each dressing', 'Qty': 3, 'Unit_Cost': 8.50, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A6243', 'BOC': 'S01', 'Desc': 'Hydrogel dressing, wound cover, sterile, pad size more than 16 sq. in. but less than or equal to 48 sq. in., without adhesive border, each dressing', 'Qty': 3, 'Unit_Cost': 12.00, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
    {'HCPCS': 'A6261', 'BOC': 'S01', 'Desc': 'Wound filler, gel/paste, per fluid ounce, not otherwise specified', 'Qty': 5, 'Unit_Cost': 7.50, 'Priority': 55, 'Source': 'LAUNCH_INVENTORY', 'Customer': ''},
])

print(f"   Added {len(missing_items)} missing items")

# Step 4: Convert missing items to DataFrame format and add to existing
print("\n[4/8] Merging with existing inventory...")

# Prepare missing items dataframe
missing_df_rows = []
for item in missing_items:
    hcpcs = item['HCPCS']
    medicare_rate = medicare_rates_dict.get(hcpcs, 0.0)
    unit_cost = item['Unit_Cost']
    qty = item['Qty']
    total_cost = unit_cost * qty

    # Calculate revenue and profit
    if medicare_rate > 0:
        total_revenue = medicare_rate * qty
        total_profit = total_revenue - total_cost
        profit_margin_pct = ((medicare_rate - unit_cost) / medicare_rate) * 100 if medicare_rate > 0 else 0
        profit_per_unit = medicare_rate - unit_cost
    else:
        total_revenue = 0
        total_profit = -total_cost  # Loss if no Medicare reimbursement
        profit_margin_pct = 0
        profit_per_unit = 0

    missing_df_rows.append({
        'HCPCS_Code': hcpcs,
        'BOC_Category': item['BOC'],
        'Description': item['Desc'],
        'Quantity': qty,
        'Estimated_Unit_Cost': unit_cost,
        'Total_Cost': total_cost,
        'Medicare_Rate': medicare_rate,
        'Priority_Score': item['Priority'],
        'Source': item['Source'],
        'Customers': item['Customer'],
        'Profit_Margin_%': profit_margin_pct,
        'Profit_Per_Unit': profit_per_unit,
        'Total_Revenue_Potential': total_revenue,
        'Total_Profit_Potential': total_profit
    })

missing_df = pd.DataFrame(missing_df_rows)
print(f"   Prepared {len(missing_df)} missing items for merge")

# Merge with existing
complete_df = pd.concat([existing_df, missing_df], ignore_index=True)

# Remove duplicates (keep first occurrence)
complete_df = complete_df.drop_duplicates(subset=['HCPCS_Code', 'Description'], keep='first')
print(f"   Total items after merge and deduplication: {len(complete_df)}")

# Step 5: Sort by priority score (highest first)
print("\n[5/8] Sorting by priority...")
complete_df = complete_df.sort_values('Priority_Score', ascending=False).reset_index(drop=True)
print(f"   Sorted {len(complete_df)} items by priority (highest first)")

# Step 6: Calculate summary statistics
print("\n[6/8] Calculating summary statistics...")

total_items = len(complete_df)
total_investment = complete_df['Total_Cost'].sum()
total_revenue_potential = complete_df['Total_Revenue_Potential'].sum()
total_profit_potential = complete_df['Total_Profit_Potential'].sum()
avg_margin = (total_profit_potential / total_revenue_potential * 100) if total_revenue_potential > 0 else 0
unique_boc_categories = complete_df['BOC_Category'].nunique()
items_with_medicare = len(complete_df[complete_df['Medicare_Rate'] > 0])

print(f"""
   Summary Statistics:
   - Total Items: {total_items}
   - Total Investment: ${total_investment:,.2f}
   - Total Revenue Potential: ${total_revenue_potential:,.2f}
   - Total Profit Potential: ${total_profit_potential:,.2f}
   - Average Margin: {avg_margin:.1f}%
   - Unique BOC Categories: {unique_boc_categories}
   - Items with Medicare Coverage: {items_with_medicare} ({items_with_medicare/total_items*100:.1f}%)
""")

# Step 7: Create Excel workbook
print("\n[7/8] Creating Excel workbook...")

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# === SHEET 1: Complete Inventory Analysis ===
ws1 = wb.create_sheet("Complete Inventory Analysis")

# Add headers
headers = ['HCPCS Code', 'BOC Category', 'Description', 'Quantity', 'Unit Cost',
           'Total Cost', 'Medicare Rate', 'Priority Score', 'Source', 'Customers',
           'Profit Margin %', 'Profit Per Unit', 'Revenue Potential', 'Profit Potential']

for col_idx, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Add data
for row_idx, (_, row) in enumerate(complete_df.iterrows(), 2):
    ws1.cell(row=row_idx, column=1, value=row['HCPCS_Code'])
    ws1.cell(row=row_idx, column=2, value=row['BOC_Category'])
    ws1.cell(row=row_idx, column=3, value=row['Description'])
    ws1.cell(row=row_idx, column=4, value=int(row['Quantity']))
    ws1.cell(row=row_idx, column=5, value=float(row['Estimated_Unit_Cost']))
    ws1.cell(row=row_idx, column=6, value=float(row['Total_Cost']))
    ws1.cell(row=row_idx, column=7, value=float(row['Medicare_Rate']) if row['Medicare_Rate'] > 0 else '')
    ws1.cell(row=row_idx, column=8, value=int(row['Priority_Score']))
    ws1.cell(row=row_idx, column=9, value=row['Source'])
    ws1.cell(row=row_idx, column=10, value=row['Customers'])
    ws1.cell(row=row_idx, column=11, value=float(row['Profit_Margin_%']))
    ws1.cell(row=row_idx, column=12, value=float(row['Profit_Per_Unit']))
    ws1.cell(row=row_idx, column=13, value=float(row['Total_Revenue_Potential']))
    ws1.cell(row=row_idx, column=14, value=float(row['Total_Profit_Potential']))

# Format columns
for col in ['E', 'F', 'G', 'L', 'M', 'N']:
    for cell in ws1[col][1:]:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'

for cell in ws1['K'][1:]:
    if cell.value and isinstance(cell.value, (int, float)):
        cell.number_format = '0.0"%"'

# Set column widths
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 60
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 14
ws1.column_dimensions['H'].width = 10
ws1.column_dimensions['I'].width = 18
ws1.column_dimensions['J'].width = 18
ws1.column_dimensions['K'].width = 12
ws1.column_dimensions['L'].width = 14
ws1.column_dimensions['M'].width = 14
ws1.column_dimensions['N'].width = 14

# Freeze header row
ws1.freeze_panes = 'A2'

print(f"   Added Sheet 1: Complete Inventory Analysis ({len(complete_df)} items)")

# === SHEET 2: BOC Category Summary ===
ws2 = wb.create_sheet("BOC Category Summary")

boc_summary = complete_df.groupby('BOC_Category').agg({
    'HCPCS_Code': 'count',
    'Total_Cost': 'sum',
    'Total_Revenue_Potential': 'sum',
    'Total_Profit_Potential': 'sum',
    'Profit_Margin_%': 'mean'
}).round(2)

boc_summary.columns = ['SKU Count', 'Total Investment', 'Revenue Potential', 'Profit Potential', 'Avg Margin %']
boc_summary = boc_summary.sort_values('Profit Potential', ascending=False)

# Add headers
headers2 = ['BOC Category', 'SKU Count', 'Total Investment', 'Revenue Potential', 'Profit Potential', 'Avg Margin %', 'ROI %']
for col_idx, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data
for row_idx, (boc_cat, row) in enumerate(boc_summary.iterrows(), 2):
    ws2.cell(row=row_idx, column=1, value=boc_cat)
    ws2.cell(row=row_idx, column=2, value=int(row['SKU Count']))
    ws2.cell(row=row_idx, column=3, value=float(row['Total Investment']))
    ws2.cell(row=row_idx, column=4, value=float(row['Revenue Potential']))
    ws2.cell(row=row_idx, column=5, value=float(row['Profit Potential']))
    ws2.cell(row=row_idx, column=6, value=float(row['Avg Margin %']))
    roi = (row['Profit Potential'] / row['Total Investment'] * 100) if row['Total Investment'] > 0 else 0
    ws2.cell(row=row_idx, column=7, value=float(roi))

# Format columns
for col in ['C', 'D', 'E']:
    for cell in ws2[col][1:]:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'

for col in ['F', 'G']:
    for cell in ws2[col][1:]:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '0.0"%"'

# Set column widths
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 12

ws2.freeze_panes = 'A2'

print(f"   Added Sheet 2: BOC Category Summary ({len(boc_summary)} categories)")

# === SHEET 3: Items Without Medicare Rates ===
ws3 = wb.create_sheet("Items Without Medicare Rates")

no_medicare_df = complete_df[complete_df['Medicare_Rate'] == 0].copy()
no_medicare_df = no_medicare_df[['HCPCS_Code', 'BOC_Category', 'Description', 'Quantity',
                                   'Estimated_Unit_Cost', 'Total_Cost', 'Source', 'Customers']]

# Add headers
headers3 = ['HCPCS Code', 'BOC Category', 'Description', 'Quantity', 'Unit Cost', 'Total Cost', 'Source', 'Customers']
for col_idx, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data
for row_idx, (_, row) in enumerate(no_medicare_df.iterrows(), 2):
    ws3.cell(row=row_idx, column=1, value=row['HCPCS_Code'])
    ws3.cell(row=row_idx, column=2, value=row['BOC_Category'])
    ws3.cell(row=row_idx, column=3, value=row['Description'])
    ws3.cell(row=row_idx, column=4, value=int(row['Quantity']))
    ws3.cell(row=row_idx, column=5, value=float(row['Estimated_Unit_Cost']))
    ws3.cell(row=row_idx, column=6, value=float(row['Total_Cost']))
    ws3.cell(row=row_idx, column=7, value=row['Source'])
    ws3.cell(row=row_idx, column=8, value=row['Customers'])

# Format columns
for col in ['E', 'F']:
    for cell in ws3[col][1:]:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'

# Set column widths
ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 60
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 12
ws3.column_dimensions['F'].width = 12
ws3.column_dimensions['G'].width = 18
ws3.column_dimensions['H'].width = 18

ws3.freeze_panes = 'A2'

print(f"   Added Sheet 3: Items Without Medicare Rates ({len(no_medicare_df)} items)")

# === SHEET 4: Customer-Specific Items ===
ws4 = wb.create_sheet("Customer Requests")

customer_df = complete_df[complete_df['Source'] == 'CUSTOMER'].copy()
customer_df = customer_df.sort_values('Priority_Score', ascending=False)
customer_df = customer_df[['HCPCS_Code', 'BOC_Category', 'Description', 'Quantity',
                            'Estimated_Unit_Cost', 'Total_Cost', 'Medicare_Rate',
                            'Priority_Score', 'Customers', 'Profit_Margin_%']]

# Add headers
headers4 = ['HCPCS Code', 'BOC Category', 'Description', 'Quantity', 'Unit Cost',
            'Total Cost', 'Medicare Rate', 'Priority', 'Customer', 'Margin %']
for col_idx, header in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data
for row_idx, (_, row) in enumerate(customer_df.iterrows(), 2):
    ws4.cell(row=row_idx, column=1, value=row['HCPCS_Code'])
    ws4.cell(row=row_idx, column=2, value=row['BOC_Category'])
    ws4.cell(row=row_idx, column=3, value=row['Description'])
    ws4.cell(row=row_idx, column=4, value=int(row['Quantity']))
    ws4.cell(row=row_idx, column=5, value=float(row['Estimated_Unit_Cost']))
    ws4.cell(row=row_idx, column=6, value=float(row['Total_Cost']))
    ws4.cell(row=row_idx, column=7, value=float(row['Medicare_Rate']) if row['Medicare_Rate'] > 0 else '')
    ws4.cell(row=row_idx, column=8, value=int(row['Priority_Score']))
    ws4.cell(row=row_idx, column=9, value=row['Customers'])
    ws4.cell(row=row_idx, column=10, value=float(row['Profit_Margin_%']))

# Format columns
for col in ['E', 'F', 'G']:
    for cell in ws4[col][1:]:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'

for cell in ws4['J'][1:]:
    if cell.value and isinstance(cell.value, (int, float)):
        cell.number_format = '0.0"%"'

# Set column widths
ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 60
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 14
ws4.column_dimensions['H'].width = 10
ws4.column_dimensions['I'].width = 18
ws4.column_dimensions['J'].width = 12

ws4.freeze_panes = 'A2'

print(f"   Added Sheet 4: Customer Requests ({len(customer_df)} items)")

# === SHEET 5: Launch Inventory by Tier ===
ws5 = wb.create_sheet("Launch Inventory by Tier")

# Tier mapping
tier_mapping = {
    'M05': 'TIER 1', 'M01': 'TIER 1', 'M06': 'TIER 1', 'M06A': 'TIER 1', 'DM02': 'TIER 1',
    'S01': 'TIER 2', 'S04': 'TIER 2',
    'PD09': 'TIER 3', 'DM29': 'TIER 3',
    'DM05': 'TIER 4', 'DM06': 'TIER 4', 'DM25': 'TIER 4',
    'R07': 'TIER 5', 'DM22': 'TIER 5', 'DM16': 'TIER 5', 'DM18': 'TIER 5',
    'DM20': 'TIER 6', 'DM08': 'TIER 6', 'DM11': 'TIER 6',
    'OR03': 'TIER 7', 'DM12': 'TIER 7', 'DM15': 'TIER 7', 'DM17': 'TIER 7', 'DM21': 'TIER 7',
    'PE03': 'TIER 8', 'PE04': 'TIER 8',
    'PD08': 'TIER 9'
}

complete_df['Tier'] = complete_df['BOC_Category'].map(tier_mapping)
tier_df = complete_df[complete_df['Tier'].notna()].copy()
tier_df = tier_df.sort_values(['Tier', 'Priority_Score'], ascending=[True, False])
tier_df_display = tier_df[['Tier', 'BOC_Category', 'HCPCS_Code', 'Description', 'Quantity',
                            'Estimated_Unit_Cost', 'Total_Cost', 'Medicare_Rate', 'Priority_Score']]

# Add headers
headers5 = ['Tier', 'BOC Category', 'HCPCS Code', 'Description', 'Quantity',
            'Unit Cost', 'Total Cost', 'Medicare Rate', 'Priority']
for col_idx, header in enumerate(headers5, 1):
    cell = ws5.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data
for row_idx, (_, row) in enumerate(tier_df_display.iterrows(), 2):
    ws5.cell(row=row_idx, column=1, value=row['Tier'])
    ws5.cell(row=row_idx, column=2, value=row['BOC_Category'])
    ws5.cell(row=row_idx, column=3, value=row['HCPCS_Code'])
    ws5.cell(row=row_idx, column=4, value=row['Description'])
    ws5.cell(row=row_idx, column=5, value=int(row['Quantity']))
    ws5.cell(row=row_idx, column=6, value=float(row['Estimated_Unit_Cost']))
    ws5.cell(row=row_idx, column=7, value=float(row['Total_Cost']))
    ws5.cell(row=row_idx, column=8, value=float(row['Medicare_Rate']) if row['Medicare_Rate'] > 0 else '')
    ws5.cell(row=row_idx, column=9, value=int(row['Priority_Score']))

# Format columns
for col in ['F', 'G', 'H']:
    for cell in ws5[col][1:]:
        if cell.value and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'

# Set column widths
ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 12
ws5.column_dimensions['C'].width = 12
ws5.column_dimensions['D'].width = 60
ws5.column_dimensions['E'].width = 10
ws5.column_dimensions['F'].width = 12
ws5.column_dimensions['G'].width = 12
ws5.column_dimensions['H'].width = 14
ws5.column_dimensions['I'].width = 10

ws5.freeze_panes = 'A2'

print(f"   Added Sheet 5: Launch Inventory by Tier ({len(tier_df_display)} items)")

# Step 8: Save workbook
print("\n[8/8] Saving workbook...")
output_filename = f"Holistic_Medical_COMPLETE_Initial_Inventory_Plan_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
wb.save(output_filename)

print(f"\n✅ SUCCESS! Workbook saved as: {output_filename}")
print("\n" + "="*80)
print("FINAL STATISTICS")
print("="*80)
print(f"Total Items: {total_items}")
print(f"Total BOC Categories: {unique_boc_categories}")
print(f"Total Investment: ${total_investment:,.2f}")
print(f"Total Revenue Potential: ${total_revenue_potential:,.2f}")
print(f"Total Profit Potential: ${total_profit_potential:,.2f}")
print(f"Average Margin: {avg_margin:.1f}%")
print(f"Medicare Coverage: {items_with_medicare}/{total_items} ({items_with_medicare/total_items*100:.1f}%)")
print("\n" + "="*80)
print("SHEETS CREATED:")
print("="*80)
print("1. Complete Inventory Analysis - All items sorted by priority")
print("2. BOC Category Summary - Category-level investment analysis")
print("3. Items Without Medicare Rates - Non-Medicare items flagged")
print("4. Customer Requests - Customer-specific items isolated")
print("5. Launch Inventory by Tier - 9-tier launch plan structure")
print("="*80)

# Save updated CSV
complete_df.to_csv('MASTER_INVENTORY_PLAN_COMPLETE_FINAL.csv', index=False)
print(f"\n📊 Also saved updated CSV as: MASTER_INVENTORY_PLAN_COMPLETE_FINAL.csv")
